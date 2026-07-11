import io, os, uuid, json, zipfile, shutil, re, secrets, textwrap
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import quote as url_quote
from datetime import datetime, timedelta
from functools import wraps

from flask import (Flask, render_template, request, jsonify,
                   send_file, session, redirect, url_for, flash)
from flask_sqlalchemy import SQLAlchemy
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
from PIL import Image, ImageDraw, ImageFont
from sqlalchemy import text

app = Flask(__name__)
app.secret_key = os.environ.get('SESSION_SECRET', 'dev-only-change-in-prod')
app.jinja_env.filters['fromjson'] = json.loads

@app.after_request
def no_cache(response):
    if 'text/html' in response.content_type:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

_db_url = os.environ.get('DATABASE_URL', 'sqlite:///catalogkit.db')
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
_engine_opts = {
    'pool_pre_ping': True,   # discard stale connections before use
    'pool_recycle':  280,    # refresh before Supabase closes idle at 5 min
}
# Only add sslmode if using PostgreSQL and the URL doesn't already set it
if _db_url.startswith('postgresql') and 'sslmode' not in _db_url:
    _engine_opts['connect_args'] = {'sslmode': 'require'}
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = _engine_opts

# ── Mail config (silently skipped if not set) ─────────────────────────────────
_mail_user = os.environ.get('MAIL_USERNAME', '')
app.config['MAIL_SERVER']         = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT']           = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS']        = os.environ.get('MAIL_USE_TLS', 'true').lower() == 'true'
app.config['MAIL_USERNAME']       = _mail_user
app.config['MAIL_PASSWORD']       = os.environ.get('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_FROM', _mail_user)

db   = SQLAlchemy(app)
mail = Mail(app)

# ── Storage (Supabase when configured, local filesystem otherwise) ─────────────
# Prefer the service-role key for server-side storage ops: this backend already
# gates uploads behind login_required, and the service key bypasses bucket RLS
# policies that would otherwise block writes made with the anon key.
_SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
_SUPABASE_KEY = (os.environ.get('SUPABASE_SERVICE_KEY')
                 or os.environ.get('SUPABASE_ANON_KEY', ''))
_supabase_client = None
if _SUPABASE_URL and _SUPABASE_KEY:
    from supabase import create_client
    _supabase_client = create_client(_SUPABASE_URL, _SUPABASE_KEY)

# Local-storage root: flask-app/static/local_storage/<bucket>/<path>
_LOCAL_STORAGE_ROOT = os.path.join(os.path.dirname(__file__), 'static', 'local_storage')

def _local_path(bucket, path):
    return os.path.join(_LOCAL_STORAGE_ROOT, bucket, path)

def _local_url(bucket, path):
    """Return a URL the browser can reach via Flask's static file serving."""
    return f'/static/local_storage/{bucket}/{path}'

def _sbucket(name):
    if not _supabase_client:
        raise RuntimeError('Supabase not configured — set SUPABASE_URL and SUPABASE_ANON_KEY.')
    return _supabase_client.storage.from_(name)

def storage_upload(bucket, path, data, content_type='image/jpeg'):
    """Returns (ok, error_detail). error_detail is the raw exception text so
    callers/UI can surface the *real* reason (e.g. bucket missing, bad key)
    instead of a generic message."""
    if _supabase_client:
        try:
            _sbucket(bucket).upload(path, data, {'content-type': content_type, 'x-upsert': 'true'})
            return True, None
        except Exception as e:
            app.logger.error('storage_upload (supabase) failed: %s', e)
            return False, str(e)
    # Local filesystem fallback
    try:
        dest = _local_path(bucket, path)
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        with open(dest, 'wb') as f:
            f.write(data if isinstance(data, bytes) else data.read())
        return True, None
    except Exception as e:
        app.logger.error('storage_upload (local) failed: %s', e)
        return False, str(e)

def storage_download(bucket, path):
    if _supabase_client:
        try:
            return _sbucket(bucket).download(path)
        except Exception:
            return None
    # Local filesystem fallback
    try:
        with open(_local_path(bucket, path), 'rb') as f:
            return f.read()
    except Exception:
        return None

def storage_public_url(bucket, path):
    if _supabase_client:
        try:
            return _sbucket(bucket).get_public_url(path)
        except Exception:
            return ''
    return _local_url(bucket, path)

def storage_delete(bucket, paths):
    if _supabase_client:
        try:
            _sbucket(bucket).remove(paths)
        except Exception:
            pass
        return
    # Local filesystem fallback
    for p in paths:
        try:
            os.remove(_local_path(bucket, p))
        except Exception:
            pass

def storage_list_prefix(bucket, prefix):
    if _supabase_client:
        try:
            items = _sbucket(bucket).list(prefix)
            return [f['name'] for f in (items or []) if isinstance(f, dict) and f.get('name')]
        except Exception:
            return []
    # Local filesystem fallback
    base = _local_path(bucket, prefix)
    if not os.path.isdir(base):
        return []
    return [f for f in os.listdir(base) if os.path.isfile(os.path.join(base, f))]

@app.template_global()
def storage_url(bucket, path):
    if not path:
        return ''
    try:
        return storage_public_url(bucket, path)
    except Exception:
        return ''

# ── Constants ─────────────────────────────────────────────────────────────────
TARGET_WIDTH  = 800
TARGET_HEIGHT = 1000
ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}

BUCKET_LOGOS  = 'logos'
BUCKET_IMAGES = 'catalog-images'

PAYMENT_INFO = {
    'contact': os.environ.get('ADMIN_WHATSAPP', '+67573817000'),
}


# ── Models ────────────────────────────────────────────────────────────────────

class User(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    email         = db.Column(db.String(255), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    name          = db.Column(db.String(255), nullable=False)
    business_name = db.Column(db.String(255), nullable=True)
    contact_person= db.Column(db.String(255), nullable=True)
    location      = db.Column(db.String(255), nullable=True)
    whatsapp      = db.Column(db.String(50),  nullable=True)
    phone         = db.Column(db.String(50),  nullable=True)
    facebook_url     = db.Column(db.String(500), nullable=True)
    payment_methods      = db.Column(db.Text, nullable=True)
    bank_account_details = db.Column(db.Text, nullable=True)
    delivery_methods     = db.Column(db.Text, nullable=True)
    reset_token      = db.Column(db.String(100), nullable=True)
    reset_token_exp  = db.Column(db.DateTime, nullable=True)
    logo_filename = db.Column(db.String(500), nullable=True)
    brand_color   = db.Column(db.String(7),   nullable=True)   # hex e.g. #7c5cfc
    pdf_layout    = db.Column(db.String(20),  default='classic')
    plan                = db.Column(db.String(20), default='free')
    plan_expires        = db.Column(db.DateTime, nullable=True)
    plan_start          = db.Column(db.DateTime, nullable=True)
    monthly_builds_used = db.Column(db.Integer, default=0)
    monthly_reset_date  = db.Column(db.Date, nullable=True)
    is_admin      = db.Column(db.Boolean, default=False)
    is_moderator  = db.Column(db.Boolean, default=False)
    is_tester     = db.Column(db.Boolean, default=False)
    is_suspended  = db.Column(db.Boolean, default=False)
    suspended_at  = db.Column(db.DateTime, nullable=True)
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    catalogs      = db.relationship('Catalog', backref='user', lazy=True, cascade='all, delete-orphan')

    @property
    def plan_label(self):
        if self.is_admin:     return 'Admin'
        if self.is_moderator: return 'Moderator'
        if self.plan == 'pro':   return 'Pro'
        if self.plan == 'basic': return 'Basic'
        return 'Free'

    @property
    def plan_css(self):
        if self.is_admin:     return 'admin'
        if self.is_moderator: return 'moderator'
        if self.plan == 'pro':   return 'pro'
        if self.plan == 'basic': return 'basic'
        return 'free'

    @property
    def monthly_limit(self):
        """Max builds per month. None = unlimited."""
        if self.is_admin or self.is_moderator: return None
        if self.plan == 'pro':   return None
        if self.plan == 'basic': return 20
        return 3

    @property
    def builds_remaining(self):
        limit = self.monthly_limit
        if limit is None: return None
        used = self.monthly_builds_used or 0
        return max(0, limit - used)

    def check_and_expire_plan(self):
        """Downgrade to free if the paid plan's expiry date has passed. Returns True if expired."""
        if self.plan != 'free' and self.plan_expires and self.plan_expires < datetime.utcnow():
            self.plan = 'free'
            self.plan_start = None
            return True
        return False

    def reset_monthly_if_needed(self):
        """Reset build count when a new month begins. Also checks plan expiry. Returns True if reset."""
        from datetime import date
        self.check_and_expire_plan()
        today = date.today()
        reset = self.monthly_reset_date
        if reset is None or today >= reset:
            self.monthly_builds_used = 0
            if today.month == 12:
                self.monthly_reset_date = date(today.year + 1, 1, 1)
            else:
                self.monthly_reset_date = date(today.year, today.month + 1, 1)
            return True
        return False

    @property
    def can_create_catalog(self):
        return True

    @property
    def catalog_count(self):
        return Catalog.query.filter_by(user_id=self.id).count()

    @property
    def profile_complete(self):
        return bool(self.business_name and self.whatsapp)


class Catalog(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    user_id       = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    name          = db.Column(db.String(255), default='My Catalog')
    pages         = db.Column(db.Text, nullable=True)   # JSON list of processed filenames
    page_count    = db.Column(db.Integer, default=0)
    pdf_downloads = db.Column(db.Integer, default=0)    # total PDFs downloaded for this catalog
    is_published  = db.Column(db.Boolean, default=False) # locked after first build
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at    = db.Column(db.DateTime, default=datetime.utcnow)

    def get_pages(self):
        """Return list of filenames only (backward compat)."""
        if not self.pages:
            return []
        data = json.loads(self.pages)
        return [p['file'] if isinstance(p, dict) else p for p in data]

    def get_page_data(self):
        """Return list of {'file': ..., 'price': ...} dicts."""
        if not self.pages:
            return []
        data = json.loads(self.pages)
        return [p if isinstance(p, dict) else {'file': p, 'price': ''} for p in data]

    def set_pages(self, pages_list):
        """Accept list of dicts {'file':..,'price':..} or plain strings."""
        self.pages      = json.dumps(pages_list)
        self.page_count = len(pages_list)
        self.updated_at = datetime.utcnow()

    @property
    def upload_prefix(self):
        return f'uploads/{self.user_id}/{self.id}'

    @property
    def processed_prefix(self):
        return f'processed/{self.user_id}/{self.id}'


class AgencyRequest(db.Model):
    """Done-For-You K50 on-site setup requests."""
    id                 = db.Column(db.Integer, primary_key=True)
    business_name      = db.Column(db.String(255), nullable=False)
    market_location    = db.Column(db.String(255), nullable=False)
    whatsapp           = db.Column(db.String(50),  nullable=False)
    preferred_datetime = db.Column(db.String(255), nullable=False)
    catalog_plan       = db.Column(db.String(20),  default='free')
    status             = db.Column(db.String(30),  default='New Request')
    submitted_at       = db.Column(db.DateTime,    default=datetime.utcnow)
    completed_at       = db.Column(db.DateTime,    nullable=True)
    user_id            = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    user               = db.relationship('User', backref=db.backref('agency_requests', lazy=True))

    @property
    def catalog_plan_label(self):
        return {'free': 'Free (K0)', 'basic': 'Basic (K20/mo)', 'pro': 'Pro (K50/mo)'}.get(self.catalog_plan, 'Free (K0)')

    @property
    def total_due(self):
        extra = {'free': 0, 'basic': 20, 'pro': 50}.get(self.catalog_plan, 0)
        return 50 + extra


class AccessLog(db.Model):
    """Records every login and logout with IP and browser info."""
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    action     = db.Column(db.String(20))           # 'login' | 'logout'
    ip_address = db.Column(db.String(45),  nullable=True)
    user_agent = db.Column(db.String(400), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user       = db.relationship('User', backref=db.backref('access_logs', lazy=True))


class ActivityLog(db.Model):
    """Records key in-app actions per user."""
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    action     = db.Column(db.String(50))           # catalog_created, pdf_downloaded, etc.
    detail     = db.Column(db.String(500), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user       = db.relationship('User', backref=db.backref('activity_logs', lazy=True))


class PaymentRequest(db.Model):
    """Manual payment approval requests for plan upgrades."""
    id             = db.Column(db.Integer, primary_key=True)
    user_id        = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    plan           = db.Column(db.String(20),  nullable=False)   # 'basic' | 'pro'
    amount         = db.Column(db.String(10),  nullable=False)   # per-month: 'K20' | 'K50'
    months_paid    = db.Column(db.Integer, default=1)            # number of months pre-paid (1–12)
    payment_method = db.Column(db.String(50),  nullable=False)   # 'cash' | 'mobile_money' | 'internet_banking'
    reference      = db.Column(db.String(255), nullable=True)    # transaction ref / receipt number
    status         = db.Column(db.String(20),  default='pending') # 'pending' | 'approved' | 'rejected'
    notes          = db.Column(db.String(500), nullable=True)    # admin notes / description
    payment_date   = db.Column(db.Date, nullable=True)           # actual date vendor paid (admin-recorded)
    submitted_at   = db.Column(db.DateTime, default=datetime.utcnow)
    resolved_at    = db.Column(db.DateTime, nullable=True)
    user           = db.relationship('User', backref=db.backref('payment_requests', lazy=True))

    @property
    def total_amount_kina(self):
        """Total payment: per-month price × months_paid."""
        per_month = int(''.join(c for c in (self.amount or '0') if c.isdigit()) or 0)
        return per_month * (self.months_paid or 1)


class AdminAuditLog(db.Model):
    """Records every change an admin/moderator makes — who did what, and when."""
    id          = db.Column(db.Integer, primary_key=True)
    admin_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    action      = db.Column(db.String(50))            # role_changed, user_suspended, announcement_sent, ...
    target_type = db.Column(db.String(30), nullable=True)   # 'user' | 'payment_request' | 'support_ticket' | 'broadcast'
    target_id   = db.Column(db.Integer, nullable=True)
    detail      = db.Column(db.String(500), nullable=True)
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    admin       = db.relationship('User', foreign_keys=[admin_id])


class SupportTicket(db.Model):
    """A vendor's support conversation with the CatalogKit team, without needing WhatsApp/email."""
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    subject    = db.Column(db.String(255), nullable=False)
    status     = db.Column(db.String(20), default='open')   # 'open' | 'in_progress' | 'resolved'
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)
    user       = db.relationship('User', backref=db.backref('support_tickets', lazy=True))

    @property
    def status_label(self):
        return {'open': 'Open', 'in_progress': 'In Progress', 'resolved': 'Resolved'}.get(self.status, 'Open')


class SupportMessage(db.Model):
    """A single message within a support ticket thread."""
    id          = db.Column(db.Integer, primary_key=True)
    ticket_id   = db.Column(db.Integer, db.ForeignKey('support_ticket.id'), nullable=False)
    from_admin  = db.Column(db.Boolean, default=False)
    author_name = db.Column(db.String(255), nullable=True)
    body        = db.Column(db.Text, nullable=False)
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    ticket      = db.relationship('SupportTicket',
                                  backref=db.backref('messages', lazy=True, order_by='SupportMessage.created_at'))


# ── Email ─────────────────────────────────────────────────────────────────────

def send_email(to, subject, body):
    if not app.config.get('MAIL_USERNAME') or not app.config.get('MAIL_PASSWORD'):
        return False
    try:
        mail.send(Message(subject, recipients=[to], body=body))
        return True
    except Exception as exc:
        app.logger.error('Email failed: %s', exc)
        return False

def admin_email():
    admin = User.query.filter_by(is_admin=True).first()
    return admin.email if admin else None

def log_access(user_id, action):
    """Record a login or logout event."""
    try:
        ip = request.headers.get('X-Forwarded-For', request.remote_addr or '')
        if ip and ',' in ip:
            ip = ip.split(',')[0].strip()
        ua = (request.user_agent.string or '')[:400]
        db.session.add(AccessLog(user_id=user_id, action=action,
                                 ip_address=ip[:45], user_agent=ua))
        db.session.commit()
    except Exception:
        pass

def log_activity(user_id, action, detail=None):
    """Record an in-app action."""
    try:
        db.session.add(ActivityLog(user_id=user_id, action=action,
                                   detail=(detail or '')[:500]))
        db.session.commit()
    except Exception:
        pass

def normalize_phone(raw):
    """Strip formatting from a phone/WhatsApp number down to its bare local digits.
    Returns None if the result is too short to be meaningful.
    Examples:  '+675 7381 7000' → '73817000'
               '07381 7000'    → '73817000'
               '73817000'      → '73817000'
    """
    if not raw:
        return None
    digits = re.sub(r'[^\d]', '', raw)
    if not digits:
        return None
    if digits.startswith('675') and len(digits) > 7:
        digits = digits[3:]
    if digits.startswith('0') and len(digits) > 7:
        digits = digits[1:]
    return digits if len(digits) >= 6 else None


def phone_in_use_by(raw_number, exclude_user_id=None):
    """Return the User already using this phone/WhatsApp number, or None.
    Pass exclude_user_id to skip the current user when updating their own profile.
    """
    norm = normalize_phone(raw_number)
    if not norm:
        return None
    q = User.query
    if exclude_user_id:
        q = q.filter(User.id != exclude_user_id)
    for u in q.all():
        if normalize_phone(u.whatsapp) == norm or normalize_phone(u.phone) == norm:
            return u
    return None


def log_admin_action(admin_id, action, target_type=None, target_id=None, detail=None):
    """Record an admin/moderator action for the security audit log."""
    try:
        db.session.add(AdminAuditLog(admin_id=admin_id, action=action,
                                     target_type=target_type, target_id=target_id,
                                     detail=(detail or '')[:500]))
        db.session.commit()
    except Exception:
        pass


# ── Image helpers ─────────────────────────────────────────────────────────────

def fit_with_padding(img, target_w, target_h, bg=(255, 255, 255)):
    """Resize image to fit inside target dimensions, pad remainder with bg colour."""
    img.thumbnail((target_w, target_h), Image.LANCZOS)
    canvas = Image.new('RGB', (target_w, target_h), bg)
    x = (target_w - img.width)  // 2
    y = (target_h - img.height) // 2
    canvas.paste(img, (x, y))
    return canvas

def allowed_file(filename):
    return os.path.splitext(filename)[1].lower() in ALLOWED_EXTENSIONS


# ── Decorators ────────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    """Allows both admins and moderators."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = db.session.get(User, session['user_id'])
        if not user or (not user.is_admin and not user.is_moderator):
            flash('Admin access required.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def full_admin_required(f):
    """Admins only — moderators cannot perform this action."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = db.session.get(User, session['user_id'])
        if not user or not user.is_admin:
            flash('Only admins can perform this action.', 'error')
            return redirect(url_for('admin'))
        return f(*args, **kwargs)
    return decorated

def current_user():
    if 'user_id' in session:
        return db.session.get(User, session['user_id'])
    return None

def get_catalog_or_404(catalog_id, user):
    catalog = db.session.get(Catalog, catalog_id)
    if not catalog or catalog.user_id != user.id:
        return None
    return catalog


# ── Profile-completion gate ────────────────────────────────────────────────────

_PROFILE_EXEMPT = {
    'profile', 'logout', 'login', 'register',
    'forgot_password', 'forgot_email', 'reset_password',
    'static', 'assisted_setup',
}

@app.before_request
def require_profile_complete():
    if 'user_id' not in session:
        return
    if request.endpoint in _PROFILE_EXEMPT or (request.endpoint or '').startswith('admin'):
        return
    user = db.session.get(User, session['user_id'])
    if user:
        # Auto-downgrade expired paid plans on every request — no cron needed
        if user.check_and_expire_plan():
            db.session.commit()
            flash('Your paid plan has expired and has been downgraded to Free. '
                  'Visit the Pricing page to renew anytime.', 'warning')
        if not user.profile_complete:
            flash('Please complete your profile before continuing.', 'info')
            return redirect(url_for('profile'))


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'user_id' in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        name             = request.form.get('name', '').strip()
        email            = request.form.get('email', '').strip().lower()
        password         = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        if not name or not email or not password or not confirm_password:
            flash('All fields are required.', 'error')
            return render_template('register.html')
        if password != confirm_password:
            flash('Passwords do not match.', 'error')
            return render_template('register.html')
        if len(password) < 6:
            flash('Password must be at least 6 characters.', 'error')
            return render_template('register.html')
        if User.query.filter_by(email=email).first():
            flash('An account with that email already exists.', 'error')
            return render_template('register.html')
        # WhatsApp duplicate check — optional field but blocks if already registered
        wa_raw = request.form.get('whatsapp', '').strip()
        if wa_raw:
            clash = phone_in_use_by(wa_raw)
            if clash:
                flash(
                    'That WhatsApp number is already linked to an existing account. '
                    'Please sign in instead, or contact us if you need help.',
                    'error'
                )
                return render_template('register.html')
        is_first = User.query.count() == 0
        user = User(name=name, email=email,
                    password_hash=generate_password_hash(password),
                    whatsapp=wa_raw or None,
                    is_admin=is_first)
        db.session.add(user)
        db.session.commit()
        session['user_id'] = user.id

        # ── Welcome email to new user ──────────────────────────────────────
        first_name = name.split()[0]
        send_email(
            email,
            'Welcome to CatalogKit! 🎉 Here\'s how to get started',
            f"""Hi {first_name},

Welcome to CatalogKit! Your account has been created and you're ready to build your first digital product catalog.

Here's how to get started in 4 easy steps:

  1. COMPLETE YOUR PROFILE
     Add your business name, WhatsApp number, and brand colour.
     This information appears on every catalog you share.
     👉 www.catalogkit.org/profile

  2. CHOOSE YOUR PLAN
     Start with the Free plan (3 builds/month at no cost),
     or upgrade to Basic (K20/month) or Pro (K50/month) for more builds.
     👉 www.catalogkit.org/pricing

  3. UPLOAD YOUR PRODUCT PHOTOS
     Drag and drop your product photos into your catalog workspace.
     Add a name and price to each product.

  4. BUILD & SHARE
     Click "Confirm & Build Catalog" and CatalogKit will generate
     your flipbook and PDF instantly — ready to share on WhatsApp!

──────────────────────────────────────
Need help getting started?

📲 WhatsApp us: +675 7381 7000
✉  Email: info@catalogkit.org
🌐 Website: www.catalogkit.org

We also offer a Done-For-You setup visit (K50 cash on site) where
our agent comes to your market stall or shop and builds your first
catalog right beside you.
👉 www.catalogkit.org/assisted-setup
──────────────────────────────────────

Thank you for joining CatalogKit. We're excited to help grow your business!

Warm regards,
The CatalogKit Team
Sapphire Consulting Services · Port Moresby, PNG
"""
        )

        # ── Notify info@catalogkit.org of new signup ───────────────────────
        send_email(
            'info@catalogkit.org',
            f'[CatalogKit] New signup — {name}',
            f"A new user has registered on CatalogKit.\n\n"
            f"Name:  {name}\n"
            f"Email: {email}\n\n"
            f"They have been sent a welcome email with getting-started steps.\n"
            f"View their account: https://www.catalogkit.org/admin\n"
        )

        flash('Account created! Please complete your profile to get started.', 'success')
        return redirect(url_for('profile'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        user     = User.query.filter_by(email=email).first()
        if not user or not check_password_hash(user.password_hash, password):
            flash('Invalid email or password.', 'error')
            return render_template('login.html')
        if user.is_suspended:
            flash('Your account has been suspended. Please contact the admin for assistance.', 'error')
            return render_template('login.html')
        session['user_id'] = user.id
        log_access(user.id, 'login')
        flash(f'Welcome back, {user.name}!', 'success')
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    if 'user_id' in session:
        log_access(session['user_id'], 'logout')
    session.clear()
    return redirect(url_for('login'))


# ── Catalog list (home) ───────────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' not in session:
        return render_template('landing.html', now=datetime.utcnow())
    user     = current_user()
    catalogs = Catalog.query.filter_by(user_id=user.id).order_by(Catalog.updated_at.desc()).all()
    return render_template('index.html', user=user, catalogs=catalogs)

@app.route('/catalog/new', methods=['POST'])
@login_required
def new_catalog():
    user = current_user()
    catalog = Catalog(user_id=user.id, name='My Catalog')
    db.session.add(catalog)
    db.session.commit()
    log_activity(user.id, 'catalog_created', catalog.name)
    return redirect(url_for('workspace', catalog_id=catalog.id))


# ── Workspace ─────────────────────────────────────────────────────────────────

@app.route('/workspace/<int:catalog_id>')
@login_required
def workspace(catalog_id):
    user    = current_user()
    catalog = get_catalog_or_404(catalog_id, user)
    if not catalog:
        flash('Catalog not found.', 'error')
        return redirect(url_for('index'))
    return render_template('workspace.html', user=user, catalog=catalog)

@app.route('/workspace/<int:catalog_id>/upload', methods=['POST'])
@login_required
def upload(catalog_id):
    user    = current_user()
    catalog = get_catalog_or_404(catalog_id, user)
    if not catalog:
        return jsonify({'error': 'Not found'}), 404
    prefix = catalog.upload_prefix
    tasks  = [(f, prefix) for f in request.files.getlist('images')
              if f and f.filename and allowed_file(f.filename)]

    if not tasks:
        return jsonify({'error': 'No valid image files received. Accepted: jpg, jpeg, png, webp, gif.'}), 400

    errors = []

    def _upload_one(args):
        f, pfx = args
        try:
            raw = f.read()
            img = Image.open(io.BytesIO(raw)).convert('RGB')
            img = fit_with_padding(img, TARGET_WIDTH, TARGET_HEIGHT)
            buf = io.BytesIO()
            img.save(buf, 'JPEG', quality=82, optimize=True)
            buf.seek(0)
            fname = f'{uuid.uuid4().hex}.jpg'
            ok, detail = storage_upload(BUCKET_IMAGES, f'{pfx}/{fname}', buf.read(), 'image/jpeg')
            if ok:
                return fname, None
            return None, f'Storage upload failed — {detail or "check Supabase credentials and bucket."}'
        except Exception as e:
            app.logger.error('upload_one failed: %s', e)
            return None, str(e)

    # Kept low (2) — Render's free-tier instance has only 512MB RAM; too much
    # concurrent Pillow decoding here was crashing/OOM-killing the single
    # gunicorn worker mid-request, which looked like a generic network failure.
    with ThreadPoolExecutor(max_workers=2) as ex:
        results = list(ex.map(_upload_one, tasks))

    uploaded = [fname for fname, err in results if fname]
    errors   = [err   for fname, err in results if err]

    if errors and not uploaded:
        # All failed — return a real error so the client shows the right message
        app.logger.error('All uploads failed. First error: %s', errors[0])
        return jsonify({'error': f'Image upload failed: {errors[0]}', 'files': []}), 500

    if uploaded:
        log_activity(user.id, 'images_uploaded',
                     f'{len(uploaded)} image(s) to "{catalog.name}"')
    return jsonify({'files': uploaded, 'failed': len(errors)})

@app.route('/workspace/<int:catalog_id>/process', methods=['POST'])
@login_required
def process(catalog_id):
    user    = current_user()
    catalog = get_catalog_or_404(catalog_id, user)
    if not catalog:
        return jsonify({'error': 'Not found'}), 404

    # ── Subscription limit check ───────────────────────────────────────────────
    if not user.is_admin and not user.is_moderator:
        user.reset_monthly_if_needed()
        limit = user.monthly_limit
        if limit is not None and (user.monthly_builds_used or 0) >= limit:
            db.session.commit()
            plan_names = {'free': 'Free (3/month)', 'basic': 'Basic (20/month)'}
            label = plan_names.get(user.plan, user.plan_label)
            return jsonify({
                'error': f"You've used all {limit} builds for this month on the {label} plan. "
                         f"Upgrade your plan or wait until next month.",
                'limit_reached': True
            }), 403

    data      = request.get_json()
    order     = data.get('order', [])
    name      = data.get('name', '').strip() or catalog.name
    prices    = data.get('prices', {})
    names     = data.get('names',  {})

    old_files = storage_list_prefix(BUCKET_IMAGES, catalog.processed_prefix)
    if old_files:
        storage_delete(BUCKET_IMAGES, [f'{catalog.processed_prefix}/{n}' for n in old_files])

    up_pfx   = catalog.upload_prefix
    proc_pfx = catalog.processed_prefix

    def _process_one(args):
        i, filename = args
        raw = storage_download(BUCKET_IMAGES, f'{up_pfx}/{filename}')
        if raw is None:
            raw = storage_download(BUCKET_IMAGES, f'{proc_pfx}/{filename}')
        if raw is None:
            return i, None, f'Missing: {filename}'
        try:
            img      = Image.open(io.BytesIO(raw)).convert('RGB')
            img      = fit_with_padding(img, TARGET_WIDTH, TARGET_HEIGHT)
            out_name = f'page_{i + 1:03d}.jpg'
            buf = io.BytesIO()
            img.save(buf, 'JPEG', quality=82, optimize=True, progressive=True)
            buf.seek(0)
            ok, detail = storage_upload(BUCKET_IMAGES, f'{proc_pfx}/{out_name}', buf.read())
            if ok:
                return i, {
                    'file':      out_name,
                    'src_file':  filename,
                    'price':     prices.get(filename, ''),
                    'item_name': names.get(filename, ''),
                }, None
            return i, None, f'Upload failed: {filename} — {detail or "unknown error"}'
        except Exception as e:
            return i, None, str(e)

    # Serial (1) — running these downloads/uploads concurrently against the
    # same Supabase client was intermittently dropping pages (some images
    # would silently fail to download/upload while others succeeded). Doing
    # this in-order trades a bit of speed for reliability, which matters more
    # for a handful of catalog images.
    with ThreadPoolExecutor(max_workers=1) as ex:
        results = sorted(ex.map(_process_one, enumerate(order)), key=lambda r: r[0])

    processed = [r[1] for r in results if r[1] is not None]
    errors    = [r[2] for r in results if r[2] is not None]

    catalog.name         = name
    catalog.is_published = True
    catalog.set_pages(processed)
    if not user.is_admin and not user.is_moderator:
        user.monthly_builds_used = (user.monthly_builds_used or 0) + 1
    db.session.commit()
    log_activity(user.id, 'catalog_published', f'{name} ({len(processed)} pages)')
    remaining = user.builds_remaining
    return jsonify({
        'processed':  [p['file'] for p in processed],
        'count':      len(processed),
        'errors':     errors,
        'builds_remaining': remaining,
    })

@app.route('/workspace/<int:catalog_id>/rename', methods=['POST'])
@login_required
def rename_catalog(catalog_id):
    user    = current_user()
    catalog = get_catalog_or_404(catalog_id, user)
    if not catalog:
        return jsonify({'error': 'Not found'}), 404
    name = (request.get_json() or {}).get('name', '').strip()
    if name:
        catalog.name       = name
        catalog.updated_at = datetime.utcnow()
        db.session.commit()
        log_activity(user.id, 'catalog_renamed', name)
    return jsonify({'name': catalog.name})

@app.route('/workspace/<int:catalog_id>/unlock', methods=['POST'])
@login_required
def unlock_catalog(catalog_id):
    """Seller has confirmed they understand editing a published catalog will
    consume another build once they re-confirm/rebuild. Unpublishes it so it
    becomes editable again, without wiping the existing pages/prices."""
    user    = current_user()
    catalog = get_catalog_or_404(catalog_id, user)
    if not catalog:
        return jsonify({'error': 'Not found'}), 404
    if not catalog.is_published:
        return jsonify({'ok': True, 'already_unlocked': True})

    if not user.is_admin and not user.is_moderator:
        user.reset_monthly_if_needed()
        limit = user.monthly_limit
        if limit is not None and (user.monthly_builds_used or 0) >= limit:
            db.session.commit()
            plan_names = {'free': 'Free (3/month)', 'basic': 'Basic (20/month)'}
            label = plan_names.get(user.plan, user.plan_label)
            return jsonify({
                'error': f"You've used all {limit} builds for this month on the {label} plan. "
                         f"Upgrade your plan or wait until next month to edit this catalog.",
                'limit_reached': True
            }), 403

    catalog.is_published = False
    db.session.commit()
    log_activity(user.id, 'catalog_unlocked', catalog.name)
    return jsonify({'ok': True})

@app.route('/workspace/<int:catalog_id>/clear', methods=['POST'])
@login_required
def clear(catalog_id):
    user    = current_user()
    catalog = get_catalog_or_404(catalog_id, user)
    if not catalog:
        return jsonify({'error': 'Not found'}), 404
    for prefix in [catalog.upload_prefix, catalog.processed_prefix]:
        files = storage_list_prefix(BUCKET_IMAGES, prefix)
        if files:
            storage_delete(BUCKET_IMAGES, [f'{prefix}/{n}' for n in files])
    catalog.set_pages([])
    catalog.is_published = False
    db.session.commit()
    return jsonify({'ok': True})


# ── Flipbook view ─────────────────────────────────────────────────────────────

@app.route('/catalog/<int:catalog_id>')
def catalog_view(catalog_id):
    # Public route — no login required so customers can view shared links
    catalog = db.session.get(Catalog, catalog_id)
    if not catalog:
        flash('Catalog not found.', 'error')
        return redirect(url_for('index'))
    owner    = db.session.get(User, catalog.user_id)
    viewer   = current_user() if 'user_id' in session else None
    is_owner = viewer is not None and viewer.id == owner.id
    # Build WhatsApp deep-link with PNG number normalisation
    wa_link = None
    if owner.whatsapp:
        digits = re.sub(r'[^\d]', '', owner.whatsapp)
        if len(digits) == 8:
            digits = '675' + digits
        elif digits.startswith('0') and len(digits) == 9:
            digits = '675' + digits[1:]
        msg = "Hi! I just viewed your flipbook catalog and would like to make an order."
        wa_link = f"https://wa.me/{digits}?text={url_quote(msg)}"
    pay_list  = list(dict.fromkeys(json.loads(owner.payment_methods  or '[]')))
    delv_list = list(dict.fromkeys(json.loads(owner.delivery_methods or '[]')))
    return render_template('catalog.html', user=owner, catalog=catalog,
                           page_data=catalog.get_page_data(),
                           wa_link=wa_link,
                           payment_methods_list=pay_list,
                           delivery_methods_list=delv_list,
                           is_owner=is_owner)

_FONT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'fonts', 'Lexend-Variable.ttf')
_FONT_BOLD = 700
_FONT_REG  = 400
_font_cache = {}

def _font(weight, size):
    """Lexend variable font at a given weight — same typeface used sitewide
    for readability (dyslexia-friendly, clear letterforms)."""
    key = (weight, size)
    cached = _font_cache.get(key)
    if cached is not None:
        return cached
    try:
        f = ImageFont.truetype(_FONT_PATH, size)
        try:
            f.set_variation_by_axes([weight])
        except Exception:
            pass
    except Exception:
        f = ImageFont.load_default()
    _font_cache[key] = f
    return f

def _hex_to_rgb(hex_color, fallback=(108, 99, 255)):
    try:
        h = hex_color.lstrip('#')
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    except Exception:
        return fallback

def _accent_rgb(user):
    """Brand colour from user profile, default purple otherwise."""
    if user.brand_color:
        return _hex_to_rgb(user.brand_color)
    return (108, 99, 255)

def _brand_stripe(draw, y0, y1, width, rgb):
    for x in range(width):
        draw.line([(x, y0), (x, y1)], fill=rgb)

def _load_logo(user, max_w, max_h):
    if not user.logo_filename:
        return None
    data = storage_download(BUCKET_LOGOS, f'{user.id}/{user.logo_filename}')
    if data is None:
        return None
    try:
        logo = Image.open(io.BytesIO(data)).convert('RGBA')
        logo.thumbnail((max_w, max_h), Image.LANCZOS)
        return logo
    except Exception:
        return None

def _paste_logo_centered(img, logo, cx, y):
    """Paste RGBA logo centred at cx, top at y; returns bottom y."""
    if not logo:
        return y
    lx = cx - logo.width // 2
    img.paste(logo, (lx, y), logo)
    return y + logo.height

def _apply_watermark(img, user):
    """Watermarks disabled — app is free to use."""
    return img

def _gradient_stripe(draw, y0, y1, width):
    stops = [(108,99,255), (200,30,120), (245,0,87), (255,109,0)]
    for x in range(width):
        t   = x / max(width - 1, 1)
        seg = t * (len(stops) - 1)
        i   = min(int(seg), len(stops) - 2)
        f   = seg - i
        c   = tuple(int(stops[i][j] + (stops[i+1][j] - stops[i][j]) * f) for j in range(3))
        draw.line([(x, y0), (x, y1)], fill=c)

def _centered_text(draw, text, x, y, font, fill):
    bb = draw.textbbox((0, 0), text, font=font)
    w  = bb[2] - bb[0]
    draw.text((x - w // 2, y), text, font=font, fill=fill)

def _wrapped_text(draw, text, cx, y, font, fill, max_width):
    words = text.split()
    line  = ''
    for word in words:
        test = (line + ' ' + word).strip()
        bb   = draw.textbbox((0, 0), test, font=font)
        if bb[2] - bb[0] > max_width and line:
            _centered_text(draw, line, cx, y, font, fill)
            y    += (bb[3] - bb[1]) + 6
            line  = word
        else:
            line = test
    if line:
        _centered_text(draw, line, cx, y, font, fill)
    return y

_wrap_measure_img = Image.new('RGB', (1, 1))
_wrap_measure_draw = ImageDraw.Draw(_wrap_measure_img)

def _wrap_lines(text, font, max_width):
    words = text.split()
    lines = []
    line  = ''
    for word in words:
        test = (line + ' ' + word).strip()
        bb   = _wrap_measure_draw.textbbox((0, 0), test, font=font)
        if bb[2] - bb[0] > max_width and line:
            lines.append(line)
            line = word
        else:
            line = test
    if line:
        lines.append(line)
    return lines

def _make_cover(catalog, user):
    W, H     = 800, 1000
    accent   = _accent_rgb(user)
    can_brand = True
    layout   = (user.pdf_layout or 'classic') if can_brand else 'classic'

    # ── Layout: Modern ────────────────────────────────────────────────────────
    if layout == 'modern':
        img  = Image.new('RGB', (W, H), (248, 249, 251))
        draw = ImageDraw.Draw(img)
        # Coloured top banner
        banner_h = 340
        banner   = Image.new('RGB', (W, banner_h), accent)
        img.paste(banner, (0, 0))
        draw = ImageDraw.Draw(img)
        # Logo inside banner
        y = 50
        logo = _load_logo(user, 200, 100) if can_brand else None
        if logo:
            y = _paste_logo_centered(img, logo, W // 2, y) + 14
        else:
            y = 80
        # Catalog name in banner (white)
        _centered_text(draw, 'PRODUCT CATALOG', W // 2, y,
                       _font(_FONT_BOLD, 12), (255, 255, 255))
        y += 26
        name_font = _font(_FONT_BOLD, 50)
        bb = draw.textbbox((0, 0), catalog.name, font=name_font)
        if bb[2] - bb[0] > W - 80:
            name_font = _font(_FONT_BOLD, 36)
        _wrapped_text(draw, catalog.name, W // 2, y, name_font, (255, 255, 255), W - 80)
        # Divider
        draw.rectangle([W // 2 - 40, banner_h + 26, W // 2 + 40, banner_h + 30], fill=accent)
        # Business name below banner
        if user.business_name:
            _centered_text(draw, user.business_name.upper(), W // 2, banner_h + 46,
                           _font(_FONT_BOLD, 20), accent)
        # Contact snippet bottom
        parts = []
        if user.whatsapp: parts.append(f'WhatsApp: {user.whatsapp}')
        if user.email:    parts.append(user.email)
        if parts:
            _centered_text(draw, '  ·  '.join(parts), W // 2, H - 40,
                           _font(_FONT_REG, 13), (150, 150, 160))
        return _apply_watermark(img, user)

    # ── Layout: Bold ──────────────────────────────────────────────────────────
    elif layout == 'bold':
        r, g, b  = accent
        dark_bg  = (max(r - 35, 0), max(g - 35, 0), max(b - 35, 0))
        img  = Image.new('RGB', (W, H), dark_bg)
        draw = ImageDraw.Draw(img)
        _brand_stripe(draw, 0, 8, W, accent)
        # Logo with white pill behind it
        y = 80
        logo = _load_logo(user, 220, 110) if can_brand else None
        if logo:
            lx = (W - logo.width) // 2 - 12
            ly = y - 10
            try:
                draw.rounded_rectangle([lx, ly, lx + logo.width + 24, ly + logo.height + 20],
                                       radius=12, fill=(255, 255, 255))
            except AttributeError:
                draw.rectangle([lx, ly, lx + logo.width + 24, ly + logo.height + 20],
                               fill=(255, 255, 255))
            y = _paste_logo_centered(img, logo, W // 2, y) + 30
            draw = ImageDraw.Draw(img)
        else:
            y = 120
        # Large catalog name
        name_font = _font(_FONT_BOLD, 60)
        bb = draw.textbbox((0, 0), catalog.name, font=name_font)
        if bb[2] - bb[0] > W - 80:
            name_font = _font(_FONT_BOLD, 44)
        _wrapped_text(draw, catalog.name, W // 2, y, name_font, (255, 255, 255), W - 80)
        # Wide rule
        draw.rectangle([W // 2 - 60, y + 120, W // 2 + 60, y + 125],
                       fill=(255, 255, 255))
        # Business name
        if user.business_name:
            _centered_text(draw, user.business_name.upper(), W // 2, y + 140,
                           _font(_FONT_REG, 22), (255, 255, 255))
        return _apply_watermark(img, user)

    # ── Layout: Classic (default) ─────────────────────────────────────────────
    else:
        img  = Image.new('RGB', (W, H), (26, 26, 46))
        draw = ImageDraw.Draw(img)
        _brand_stripe(draw, 0, 6, W, accent)
        # Circle accent
        for rv in range(180, 0, -20):
            draw.ellipse([W - rv - 30, H - rv - 30, W - 30 + rv, H - 30 + rv],
                         outline=accent)
        _centered_text(draw, 'PRODUCT CATALOG', W // 2, 90,
                       _font(_FONT_BOLD, 13), (200, 200, 220))
        # Logo above catalog name for Growth users
        y = H // 2 - 120
        logo = _load_logo(user, 200, 90) if can_brand else None
        if logo:
            y = _paste_logo_centered(img, logo, W // 2, y) + 12
            draw = ImageDraw.Draw(img)
        else:
            y = H // 2 - 70
        name_font = _font(_FONT_BOLD, 52)
        bb = draw.textbbox((0, 0), catalog.name, font=name_font)
        if bb[2] - bb[0] > W - 80:
            name_font = _font(_FONT_BOLD, 36)
        _wrapped_text(draw, catalog.name, W // 2, y, name_font, (255, 255, 255), W - 80)
        draw.rectangle([W // 2 - 32, H // 2 + 10, W // 2 + 32, H // 2 + 14], fill=accent)
        if user.business_name:
            _centered_text(draw, user.business_name.upper(), W // 2, H // 2 + 34,
                           _font(_FONT_REG, 18), (200, 200, 220))
        return _apply_watermark(img, user)

def _make_product_page(item, catalog, user):
    W, H  = 800, 1000
    fname = item['file'] if isinstance(item, dict) else item
    iname = (item.get('item_name') or '') if isinstance(item, dict) else ''
    price = (item.get('price') or '')    if isinstance(item, dict) else ''
    raw   = storage_download(BUCKET_IMAGES, f'{catalog.processed_prefix}/{fname}')
    DARK   = (15, 15, 30)
    TXT    = (210, 210, 230)
    accent = _accent_rgb(user)
    # Reserve space for the bars so the photo is placed *between* them,
    # never underneath — matches the web flipbook layout.
    top_h    = 60 if iname else 28
    bottom_h = 56 if price else 28
    avail_h  = H - top_h - bottom_h
    pg = Image.new('RGB', (W, H), (255, 255, 255))
    if raw:
        photo = Image.open(io.BytesIO(raw)).convert('RGB')
        pw, ph = photo.size
        scale  = min(W / pw, avail_h / ph)
        new_w, new_h = max(1, int(pw * scale)), max(1, int(ph * scale))
        photo  = photo.resize((new_w, new_h), Image.LANCZOS)
        px = (W - new_w) // 2
        py = top_h + (avail_h - new_h) // 2
        pg.paste(photo, (px, py))
    draw = ImageDraw.Draw(pg)
    # header bar
    draw.rectangle([0, 0, W, 28], fill=DARK)
    _brand_stripe(draw, 0, 4, W, accent)
    header = f"Product Catalog  —  {catalog.name}"
    _centered_text(draw, header, W // 2, 7, _font(_FONT_REG, 13), TXT)
    # item name bar
    if iname:
        draw.rectangle([0, 28, W, 60], fill=(20, 20, 40))
        _centered_text(draw, iname, W // 2, 35, _font(_FONT_BOLD, 18), (255, 255, 255))
    # price bar
    if price:
        draw.rectangle([0, H - 56, W, H - 28], fill=(20, 20, 40))
        _centered_text(draw, price, W // 2, H - 50, _font(_FONT_BOLD, 22), (255, 255, 255))
    # footer bar
    draw.rectangle([0, H - 28, W, H], fill=DARK)
    parts = []
    if user.business_name: parts.append(user.business_name)
    if user.email:          parts.append(user.email)
    if user.whatsapp:       parts.append(f"WhatsApp: {user.whatsapp}")
    _centered_text(draw, '  |  '.join(parts), W // 2, H - 22,
                   _font(_FONT_REG, 11), (150, 150, 170))
    return pg

def _make_back_cover(catalog, user):
    W, H   = 800, 1000
    accent = _accent_rgb(user)
    img    = Image.new('RGB', (W, H), (255, 255, 255))
    draw   = ImageDraw.Draw(img)
    _brand_stripe(draw, 0, 5, W, accent)
    y = 38
    if user.business_name:
        _centered_text(draw, user.business_name.upper(), W // 2, y,
                       _font(_FONT_BOLD, 13), (156, 163, 175))
        y += 32
    _centered_text(draw, 'Thank you for supporting', W // 2, y,
                   _font(_FONT_BOLD, 34), (15, 23, 42))
    y += 44
    _centered_text(draw, 'a local PNG SME!', W // 2, y,
                   _font(_FONT_BOLD, 34), (15, 23, 42))
    y += 44
    _centered_text(draw, 'Got questions or ready to order? Follow the simple steps below.',
                   W // 2, y, _font(_FONT_REG, 16), (100, 116, 139))
    y += 36
    # steps box
    steps = [
        ('1', 'Browse & Choose', 'Pick your favourite items from this catalog.'),
        ('2', 'Contact Us',      'Send us a message on WhatsApp with your order details.'),
        ('3', 'Confirm & Pay',   'We confirm your order and agree on payment & delivery.'),
    ]
    pad  = 60
    bx0, bx1 = pad, W - pad
    bh   = 36 + len(steps) * 62
    try:
        draw.rounded_rectangle([bx0, y, bx1, y + bh], radius=10,
                                fill=(248, 249, 251), outline=(226, 232, 240))
    except AttributeError:
        draw.rectangle([bx0, y, bx1, y + bh], fill=(248, 249, 251), outline=(226, 232, 240))
    sy = y + 18
    for num, title, desc in steps:
        draw.ellipse([bx0 + 12, sy, bx0 + 38, sy + 26], fill=(108, 99, 255))
        _centered_text(draw, num, bx0 + 25, sy + 5, _font(_FONT_BOLD, 13), (255, 255, 255))
        draw.text((bx0 + 50, sy + 2),  title, font=_font(_FONT_BOLD, 15), fill=(15, 23, 42))
        draw.text((bx0 + 50, sy + 20), desc,  font=_font(_FONT_REG, 12),  fill=(100, 116, 139))
        sy += 62
    y = sy + 22
    # WhatsApp button
    if user.whatsapp:
        try:
            draw.rounded_rectangle([pad, y, W - pad, y + 44], radius=8, fill=(37, 211, 102))
        except AttributeError:
            draw.rectangle([pad, y, W - pad, y + 44], fill=(37, 211, 102))
        _centered_text(draw, f"WhatsApp Us: {user.whatsapp}", W // 2, y + 12,
                       _font(_FONT_BOLD, 17), (255, 255, 255))
        y += 58
    # payment / delivery info — per-vendor selections, matches the digital flipbook
    pay_items  = json.loads(user.payment_methods  or '[]')
    delv_items = json.loads(user.delivery_methods or '[]')
    PAYMENT_DELIVERY_INFO = [
        ('PAYMENT OPTIONS', pay_items, (237, 233, 254), (91, 33, 182)),
        ('DELIVERY & COLLECTION', delv_items, (220, 252, 231), (21, 128, 61)),
    ]
    col_w = (W - 120 - 10) // 2
    if pay_items or delv_items:
        for col_idx, (label, items, badge_fill, badge_txt) in enumerate(PAYMENT_DELIVERY_INFO):
            bx = pad + col_idx * (col_w + 10)
            wrapped = []
            for it in items:
                wrapped.extend(_wrap_lines(it, _font(_FONT_REG, 9), col_w - 20))
            if not wrapped:
                wrapped = ['Not specified']
            bh2 = 28 + len(wrapped) * 14 + 10
            try:
                draw.rounded_rectangle([bx, y, bx + col_w, y + bh2], radius=7,
                                        fill=(248, 249, 251), outline=(226, 232, 240))
            except AttributeError:
                draw.rectangle([bx, y, bx + col_w, y + bh2], fill=(248, 249, 251), outline=(226, 232, 240))
            draw.text((bx + 10, y + 8), label, font=_font(_FONT_BOLD, 11), fill=(100, 116, 139))
            my = y + 26
            for line in wrapped:
                draw.text((bx + 10, my), line, font=_font(_FONT_REG, 9), fill=(71, 85, 105))
                my += 14
        def _col_height(items):
            wrapped = []
            for it in items:
                wrapped.extend(_wrap_lines(it, _font(_FONT_REG, 9), col_w - 20))
            if not wrapped:
                wrapped = ['x']
            return 28 + len(wrapped) * 14 + 10
        y = y + max(_col_height(pay_items), _col_height(delv_items)) + 16
    # footer links
    fy = H - 52
    draw.line([(pad, fy), (W - pad, fy)], fill=(241, 245, 249), width=1)
    fy += 8
    if user.facebook_url:
        draw.text((pad, fy), f"fb  {user.facebook_url}", font=_font(_FONT_BOLD, 12), fill=(24, 119, 242))
        fy += 20
    if user.email:
        draw.text((pad, fy), f"✉  {user.email}", font=_font(_FONT_REG, 11), fill=(148, 163, 184))
    return img

def generate_catalog_pdf(catalog, user):
    page_data = catalog.get_page_data()
    pages     = [_make_cover(catalog, user)]
    for item in page_data:
        pages.append(_make_product_page(item, catalog, user))
    pages.append(_make_back_cover(catalog, user))
    buf = io.BytesIO()
    pages[0].save(buf, format='PDF', save_all=True, append_images=pages[1:], resolution=96)
    buf.seek(0)
    return buf


@app.route('/catalog/<int:catalog_id>/download')
@login_required
def download_catalog(catalog_id):
    user    = current_user()
    catalog = get_catalog_or_404(catalog_id, user)
    if not catalog:
        return jsonify({'error': 'Not found'}), 404
    if not catalog.get_pages():
        flash('No images in this catalog yet.', 'error')
        return redirect(url_for('workspace', catalog_id=catalog_id))
    buf = generate_catalog_pdf(catalog, user)
    log_activity(user.id, 'pdf_downloaded', catalog.name)
    safe_name = ''.join(c for c in catalog.name if c.isalnum() or c in ' _-')[:40].strip()
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name=f'{safe_name or "catalog"}.pdf')

@app.route('/catalog/<int:catalog_id>/delete', methods=['POST'])
@login_required
def delete_catalog(catalog_id):
    user    = current_user()
    catalog = get_catalog_or_404(catalog_id, user)
    if not catalog:
        flash('Catalog not found.', 'error')
        return redirect(url_for('index'))
    name = catalog.name
    for prefix in [catalog.upload_prefix, catalog.processed_prefix]:
        files = storage_list_prefix(BUCKET_IMAGES, prefix)
        if files:
            storage_delete(BUCKET_IMAGES, [f'{prefix}/{n}' for n in files])
    db.session.delete(catalog)
    db.session.commit()
    log_activity(user.id, 'catalog_deleted', name)
    flash(f'"{name}" deleted.', 'success')
    return redirect(url_for('index'))


# ── Dashboard (plan info) ─────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    user = current_user()
    user.reset_monthly_if_needed()
    db.session.commit()
    plan_limits = {'free': 3, 'basic': 20, 'pro': None}
    return render_template('dashboard.html', user=user,
                           catalogs=sorted(user.catalogs, key=lambda c: c.updated_at, reverse=True),
                           now=datetime.utcnow(),
                           plan_limits=plan_limits)


# ── Forgot / Reset Password ───────────────────────────────────────────────────

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    admin_whatsapp = PAYMENT_INFO.get('contact', '')
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        user  = User.query.filter_by(email=email).first()
        if user:
            token = secrets.token_urlsafe(32)
            user.reset_token     = token
            user.reset_token_exp = datetime.utcnow() + timedelta(hours=1)
            db.session.commit()
            reset_url = request.host_url.rstrip('/') + f'/reset-password/{token}'
            sent = send_email(
                user.email,
                'CatalogKit — Reset your password',
                f'Hi {user.name},\n\nClick the link below to reset your password. '
                f'It expires in 1 hour.\n\n{reset_url}\n\n'
                f'If you did not request this, you can ignore this email.'
            )
            return render_template('forgot_password.html',
                                   sent=sent,
                                   reset_url=(None if sent else reset_url),
                                   mail_configured=sent,
                                   admin_whatsapp=admin_whatsapp)
        return render_template('forgot_password.html', sent=True, reset_url=None,
                               mail_configured=True, admin_whatsapp=admin_whatsapp)
    return render_template('forgot_password.html', sent=False, reset_url=None,
                           mail_configured=None, admin_whatsapp=admin_whatsapp)


@app.route('/forgot-email', methods=['GET', 'POST'])
def forgot_email():
    """Let users look up which email they registered with using their phone/WhatsApp."""
    admin_whatsapp = PAYMENT_INFO.get('contact', '')
    found_email = None
    not_found   = False
    if request.method == 'POST':
        phone_raw = request.form.get('phone', '').strip()
        # Normalise: strip non-digits for comparison
        digits = re.sub(r'[^\d]', '', phone_raw)
        user = None
        # Try exact matches on whatsapp or phone columns
        for candidate in User.query.all():
            for field in [candidate.whatsapp, candidate.phone]:
                if field and re.sub(r'[^\d]', '', field) == digits:
                    user = candidate
                    break
            if user:
                break
        if user:
            # Partially mask the email for privacy: j***@gmail.com
            parts  = user.email.split('@')
            local  = parts[0]
            masked = local[0] + '***' + (local[-1] if len(local) > 1 else '') + '@' + parts[1]
            found_email = masked
        else:
            not_found = True
    return render_template('forgot_email.html',
                           found_email=found_email,
                           not_found=not_found,
                           admin_whatsapp=admin_whatsapp)


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    user = User.query.filter_by(reset_token=token).first()
    if not user or not user.reset_token_exp or user.reset_token_exp < datetime.utcnow():
        flash('This reset link is invalid or has expired. Please request a new one.', 'error')
        return redirect(url_for('forgot_password'))
    if request.method == 'POST':
        pw  = request.form.get('password', '')
        pw2 = request.form.get('password2', '')
        if len(pw) < 6:
            return render_template('reset_password.html', token=token, error='Password must be at least 6 characters.')
        if pw != pw2:
            return render_template('reset_password.html', token=token, error='Passwords do not match.')
        user.password_hash  = generate_password_hash(pw)
        user.reset_token    = None
        user.reset_token_exp = None
        db.session.commit()
        flash('Your password has been reset. Please sign in.', 'success')
        return redirect(url_for('login'))
    return render_template('reset_password.html', token=token, error=None)


# ── Profile ───────────────────────────────────────────────────────────────────

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    user = current_user()
    if request.method == 'POST':
        was_incomplete = not user.profile_complete
        user.business_name  = request.form.get('business_name', '').strip() or None
        user.contact_person = request.form.get('contact_person', '').strip() or None
        user.location       = request.form.get('location', '').strip() or None
        new_whatsapp = request.form.get('whatsapp', '').strip() or None
        new_phone    = request.form.get('phone', '').strip() or None
        # Duplicate phone/WhatsApp check — skip if the number belongs to this user already
        for raw in filter(None, [new_whatsapp, new_phone]):
            clash = phone_in_use_by(raw, exclude_user_id=user.id)
            if clash:
                flash(
                    f'That number ({raw}) is already linked to another account. '
                    'Each account must use a unique phone and WhatsApp number.',
                    'error'
                )
                return render_template('profile.html', user=user)
        user.whatsapp = new_whatsapp
        user.phone    = new_phone
        user.email          = request.form.get('email', '').strip().lower() or user.email
        user.facebook_url   = request.form.get('facebook_url', '').strip() or None
        pay  = request.form.getlist('payment_methods')
        delv = request.form.getlist('delivery_methods')
        user.payment_methods      = json.dumps(pay)  if pay  else None
        user.delivery_methods     = json.dumps(delv) if delv else None
        user.bank_account_details = request.form.get('bank_account_details', '').strip() or None
        log_activity(user.id, 'profile_updated', 'Profile information updated')
        # ── Branding (free for all users) ─────────────────────────────────────
        if True:
            color = request.form.get('brand_color', '').strip()
            if color and color.startswith('#') and len(color) == 7:
                user.brand_color = color
            layout = request.form.get('pdf_layout', 'classic')
            if layout in ('classic', 'modern', 'bold'):
                user.pdf_layout = layout
            # Logo upload
            logo_file = request.files.get('logo')
            if logo_file and logo_file.filename:
                ext = os.path.splitext(logo_file.filename)[1].lower()
                if ext in {'.jpg', '.jpeg', '.png', '.webp'}:
                    if user.logo_filename:
                        storage_delete(BUCKET_LOGOS, [f'{user.id}/{user.logo_filename}'])
                    fname = f"logo_{user.id}_{uuid.uuid4().hex[:8]}{ext}"
                    ct    = 'image/jpeg' if ext in ('.jpg', '.jpeg') else f'image/{ext.lstrip(".")}'
                    ok, detail = storage_upload(BUCKET_LOGOS, f'{user.id}/{fname}', logo_file.read(), ct)
                    if ok:
                        user.logo_filename = fname
                    else:
                        flash(f'Logo upload failed — {detail or "please try again."}', 'error')
                else:
                    flash('Logo must be JPG, PNG, or WebP.', 'error')
        db.session.commit()
        # First-time profile completion → go straight to a new catalog
        if was_incomplete and user.profile_complete:
            return redirect(url_for('choose_plan'))
        flash('Profile updated!', 'success')
        return redirect(url_for('profile'))
    return render_template('profile.html', user=user)


# ── Plan selection & payment approval ────────────────────────────────────────

PLAN_AMOUNTS = {'basic': 'K20', 'pro': 'K50'}
PAYMENT_METHOD_LABELS = {
    'mobile_money':      'Mobile Money (MiCash / BSP Kina)',
    'internet_banking':  'Internet Banking',
    'cash':              'Cash',
}

@app.route('/choose-plan', methods=['GET', 'POST'])
@login_required
def choose_plan():
    user = current_user()
    if request.method == 'POST':
        plan = request.form.get('plan', 'free')
        if plan == 'free':
            # Free plan — create first catalog immediately
            catalog = Catalog(user_id=user.id, name='My Catalog')
            db.session.add(catalog)
            db.session.commit()
            log_activity(user.id, 'catalog_created', catalog.name)
            flash('Welcome to CatalogKit! Your catalog is ready.', 'success')
            return redirect(url_for('workspace', catalog_id=catalog.id))
        if plan not in PLAN_AMOUNTS:
            flash('Invalid plan selected.', 'error')
            return redirect(url_for('choose_plan'))
        # Paid plan — show payment details form
        return render_template('choose_plan.html', user=user,
                               show_payment=True, plan=plan,
                               amount=PLAN_AMOUNTS[plan])
    return render_template('choose_plan.html', user=user,
                           show_payment=False, plan=None, amount=None)


@app.route('/payment-request', methods=['POST'])
@login_required
def payment_request():
    user   = current_user()
    plan   = request.form.get('plan', '')
    method = request.form.get('payment_method', '')
    ref    = request.form.get('reference', '').strip()

    if plan not in PLAN_AMOUNTS or method not in PAYMENT_METHOD_LABELS:
        flash('Invalid payment details.', 'error')
        return redirect(url_for('choose_plan'))

    try:
        months_paid = int(request.form.get('months_paid', '1') or 1)
    except (ValueError, TypeError):
        months_paid = 1
    months_paid = max(1, min(12, months_paid))

    per_month_str = PLAN_AMOUNTS[plan]
    per_month_int = int(''.join(c for c in per_month_str if c.isdigit()) or 0)
    total_kina    = per_month_int * months_paid

    pr = PaymentRequest(
        user_id        = user.id,
        plan           = plan,
        amount         = per_month_str,
        months_paid    = months_paid,
        payment_method = method,
        reference      = ref or None,
    )
    db.session.add(pr)
    db.session.commit()
    log_activity(user.id, 'payment_requested',
                 f'{plan} plan × {months_paid} month(s) = K{total_kina} via {method}')

    method_label = PAYMENT_METHOD_LABELS[method]
    months_label = f'{months_paid} month' + ('' if months_paid == 1 else 's')
    body = (
        f"New plan upgrade request — please verify and approve.\n\n"
        f"User:           {user.name} ({user.email})\n"
        f"Business:       {user.business_name or '(not set)'}\n"
        f"WhatsApp:       {user.whatsapp or '(not set)'}\n"
        f"Plan requested: {plan.title()} ({per_month_str}/month × {months_label})\n"
        f"Total amount:   K{total_kina}\n"
        f"Payment method: {method_label}\n"
        f"Reference/Ref#: {ref or '(not provided)'}\n\n"
        f"To approve, go to: https://www.catalogkit.org/admin\n"
        f"Request ID: #{pr.id}\n"
    )
    send_email('info@catalogkit.org',
               f'[CatalogKit] Payment #{pr.id} — {user.name} → {plan.title()} K{total_kina} ({months_label})',
               body)

    return redirect(url_for('payment_pending'))


@app.route('/payment-pending')
@login_required
def payment_pending():
    user = current_user()
    # Find most recent pending request for this user
    pr = PaymentRequest.query.filter_by(user_id=user.id, status='pending')\
                             .order_by(PaymentRequest.submitted_at.desc()).first()
    return render_template('payment_pending.html', user=user, pr=pr)


@app.route('/admin/payment/<int:pr_id>/approve', methods=['POST'])
@admin_required
def admin_approve_payment(pr_id):
    pr = db.session.get(PaymentRequest, pr_id)
    if not pr:
        flash('Request not found.', 'error')
        return redirect(url_for('admin'))
    pr.status      = 'approved'
    pr.resolved_at = datetime.utcnow()
    pr.notes       = request.form.get('notes', '').strip() or None

    user = db.session.get(User, pr.user_id)
    if user:
        months_paid   = pr.months_paid or 1
        per_month_int = int(''.join(c for c in (pr.amount or '0') if c.isdigit()) or 0)
        total_kina    = per_month_int * months_paid
        months_label  = f'{months_paid} month' + ('' if months_paid == 1 else 's')
        now_dt        = datetime.utcnow()

        # If the user is on the same plan and it hasn't expired yet, extend from current expiry.
        # Otherwise start fresh from today.
        if user.plan == pr.plan and user.plan_expires and user.plan_expires > now_dt:
            new_expires  = user.plan_expires + timedelta(days=30 * months_paid)
            plan_message = f'extended by {months_label} (now expires {new_expires.strftime("%d %b %Y")})'
        else:
            user.plan_start = now_dt
            new_expires     = now_dt + timedelta(days=30 * months_paid)
            plan_message    = f'activated for {months_label} (expires {new_expires.strftime("%d %b %Y")})'

        user.plan         = pr.plan
        user.plan_expires = new_expires
        # Reset the monthly build counter immediately on activation
        user.monthly_builds_used = 0
        today = now_dt.date()
        if today.month == 12:
            user.monthly_reset_date = date(today.year + 1, 1, 1)
        else:
            user.monthly_reset_date = date(today.year, today.month + 1, 1)

        if not user.catalogs:
            catalog = Catalog(user_id=user.id, name='My Catalog')
            db.session.add(catalog)
        db.session.commit()
        log_activity(user.id, 'plan_upgraded',
                     f'{pr.plan.title()} plan {plan_message} (payment #{pr.id} approved, K{total_kina} total)')
        log_admin_action(session['user_id'], 'payment_approved', 'payment_request', pr.id,
                         f'Approved {pr.plan} plan K{total_kina} ({months_label}) for {user.name} ({user.email})')
        send_email(
            user.email,
            f'[CatalogKit] Your {pr.plan.title()} plan is now active!',
            f"Hi {user.name},\n\n"
            f"We've confirmed your payment of K{total_kina} ({pr.amount}/month × {months_label}).\n\n"
            f"Your {pr.plan.title()} plan is {plan_message}.\n\n"
            f"Log in to start building your catalogs: https://www.catalogkit.org\n\n"
            f"Thank you for supporting CatalogKit!\n"
            f"— The CatalogKit Team"
        )
    else:
        db.session.commit()
    flash(f'Payment #{pr_id} approved — {pr.plan.title()} plan activated.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/payment/<int:pr_id>/reject', methods=['POST'])
@admin_required
def admin_reject_payment(pr_id):
    pr = db.session.get(PaymentRequest, pr_id)
    if not pr:
        flash('Request not found.', 'error')
        return redirect(url_for('admin'))
    pr.status      = 'rejected'
    pr.resolved_at = datetime.utcnow()
    pr.notes       = request.form.get('notes', '').strip() or None
    db.session.commit()
    user = db.session.get(User, pr.user_id)
    log_admin_action(session['user_id'], 'payment_rejected', 'payment_request', pr.id,
                     f'Rejected payment #{pr.id}' + (f' for {user.name}' if user else ''))
    if user:
        log_activity(user.id, 'payment_rejected', f'Payment #{pr.id} rejected')
        send_email(
            user.email,
            '[CatalogKit] Payment not confirmed — please contact us',
            f"Hi {user.name},\n\n"
            f"We were unable to confirm your payment for the {pr.plan.title()} plan.\n\n"
            f"Please message us on WhatsApp: +675 7381 7000\n"
            f"or email: info@catalogkit.org\n\n"
            f"— The CatalogKit Team"
        )
    flash(f'Payment #{pr_id} rejected.', 'success')
    return redirect(url_for('admin'))


# ── Admin billing management ───────────────────────────────────────────────────

@app.route('/admin/billing')
@admin_required
def admin_billing():
    now_dt = datetime.utcnow()
    seven_days_out  = now_dt + timedelta(days=7)
    thirty_days_out = now_dt + timedelta(days=30)

    # Active paid subscribers, sorted soonest-expiring first
    active_paid = User.query.filter(
        User.plan.in_(['basic', 'pro']),
        User.plan_expires > now_dt,
        User.is_admin == False,
        User.is_moderator == False,
    ).order_by(User.plan_expires).all()

    expiring_urgent = [u for u in active_paid if u.plan_expires <= seven_days_out]
    expiring_soon   = [u for u in active_paid if seven_days_out < u.plan_expires <= thirty_days_out]
    expiring_later  = [u for u in active_paid if u.plan_expires > thirty_days_out]

    # Vendors who once had a paid plan (approved payment exists) but are now on free
    lapsed_ids = {pr.user_id for pr in PaymentRequest.query.filter_by(status='approved').all()}
    lapsed = [u for u in User.query.filter(
        User.id.in_(lapsed_ids), User.plan == 'free',
        User.is_admin == False, User.is_suspended == False
    ).order_by(User.name).all()]

    # Pending payment requests submitted by vendors via the website
    pending_prs = PaymentRequest.query.filter_by(status='pending')\
                               .order_by(PaymentRequest.submitted_at).all()

    # Admin-logged payments awaiting receipt confirmation
    pending_confirmation_prs = PaymentRequest.query.filter_by(status='pending_confirmation')\
                               .order_by(PaymentRequest.submitted_at.desc()).all()

    # Full approved payment history, newest first
    approved_prs = PaymentRequest.query.filter_by(status='approved')\
                                 .order_by(PaymentRequest.resolved_at.desc()).all()

    # Revenue totals
    total_received = sum(p.total_amount_kina for p in approved_prs)
    pending_kina   = sum(
        int(''.join(c for c in (p.amount or '0') if c.isdigit()) or 0) * (p.months_paid or 1)
        for p in pending_prs + pending_confirmation_prs
    )

    # Per-user last payment dict  {user_id: PaymentRequest}
    last_payment = {}
    for pr in approved_prs:
        if pr.user_id not in last_payment:
            last_payment[pr.user_id] = pr

    # All active (non-suspended) vendors for the Record Payment modal
    all_vendors = User.query.filter_by(is_admin=False, is_moderator=False, is_suspended=False)\
                  .order_by(User.name).all()

    pending_count = len(pending_prs) + len(pending_confirmation_prs)
    return render_template(
        'admin_billing.html',
        admin_active='billing',
        active_paid=active_paid,
        expiring_urgent=expiring_urgent,
        expiring_soon=expiring_soon,
        expiring_later=expiring_later,
        lapsed=lapsed,
        pending_prs=pending_prs,
        pending_confirmation_prs=pending_confirmation_prs,
        approved_prs=approved_prs,
        total_received=total_received,
        pending_kina=pending_kina,
        last_payment=last_payment,
        all_vendors=all_vendors,
        now=now_dt,
        pending_count=pending_count,
    )


@app.route('/admin/user/<int:uid>/set-plan', methods=['POST'])
@admin_required
def admin_set_plan(uid):
    """Admin directly sets a user's plan and duration — for in-person/cash payments."""
    target = db.session.get(User, uid)
    if not target:
        flash('User not found.', 'error')
        return redirect(url_for('admin_billing'))

    new_plan   = request.form.get('plan', 'free')
    months     = int(request.form.get('months', '1') or 1)
    months     = max(1, min(24, months))
    notes      = request.form.get('notes', '').strip()
    amount_str = request.form.get('amount', '').strip()  # optional manual amount override
    now_dt     = datetime.utcnow()

    plan_prices = {'free': 0, 'basic': 20, 'pro': 50}

    if new_plan == 'free':
        target.plan         = 'free'
        target.plan_expires = None
        target.plan_start   = None
        db.session.commit()
        log_admin_action(session['user_id'], 'plan_set_free', 'user', uid,
                         f'Downgraded {target.name} to Free{(" — " + notes) if notes else ""}')
        flash(f'{target.name} downgraded to Free plan.', 'success')
        return redirect(url_for('admin_billing'))

    # Extend from current expiry if on same plan and not yet expired
    if target.plan == new_plan and target.plan_expires and target.plan_expires > now_dt:
        new_expires  = target.plan_expires + timedelta(days=30 * months)
        plan_message = f'extended by {months} month(s) → expires {new_expires.strftime("%d %b %Y")}'
    else:
        target.plan_start = now_dt
        new_expires       = now_dt + timedelta(days=30 * months)
        plan_message      = f'set to {new_plan.title()} for {months} month(s) → expires {new_expires.strftime("%d %b %Y")}'

    target.plan         = new_plan
    target.plan_expires = new_expires
    target.monthly_builds_used = 0
    today = now_dt.date()
    target.monthly_reset_date = date(today.year + 1, 1, 1) if today.month == 12 \
                                else date(today.year, today.month + 1, 1)

    # Record as a payment request so it shows in payment history
    per_month = int(amount_str.replace('K','')) if amount_str else plan_prices.get(new_plan, 0)
    pr = PaymentRequest(
        user_id        = uid,
        plan           = new_plan,
        amount         = f'K{per_month}',
        months_paid    = months,
        payment_method = 'admin_override',
        reference      = notes or 'Manual admin adjustment',
        status         = 'approved',
        submitted_at   = now_dt,
        resolved_at    = now_dt,
        notes          = notes or 'Set directly by admin',
    )
    db.session.add(pr)
    db.session.commit()

    log_admin_action(session['user_id'], 'plan_set', 'user', uid,
                     f'{target.name}: {plan_message}{(" — " + notes) if notes else ""}')
    log_activity(uid, 'plan_upgraded',
                 f'{new_plan.title()} plan {plan_message} (admin adjustment)')

    send_email(
        target.email,
        f'[CatalogKit] Your {new_plan.title()} plan is now active!',
        f"Hi {target.name},\n\n"
        f"Your {new_plan.title()} plan has been {plan_message}.\n\n"
        f"Log in to continue building your catalogs: https://www.catalogkit.org\n\n"
        f"Thank you!\n— The CatalogKit Team"
    )

    flash(f'{target.name} → {plan_message}.', 'success')
    redirect_to = request.form.get('redirect_to', 'billing')
    return redirect(url_for('admin_billing') if redirect_to == 'billing' else url_for('admin'))


@app.route('/admin/record-payment', methods=['POST'])
@admin_required
def admin_record_payment():
    """Admin manually logs a payment received via WhatsApp or email."""
    uid         = request.form.get('user_id', type=int)
    plan        = request.form.get('plan', 'basic')
    months      = max(1, min(24, int(request.form.get('months', '1') or 1)))
    method      = request.form.get('payment_method', 'cash')
    reference   = request.form.get('reference', '').strip() or None
    notes       = request.form.get('notes', '').strip() or None
    activate_now = request.form.get('activate_now') == '1'
    payment_date_str = request.form.get('payment_date', '').strip()

    target = db.session.get(User, uid)
    if not target:
        flash('Vendor not found.', 'error')
        return redirect(url_for('admin_billing'))

    payment_date = None
    if payment_date_str:
        try:
            from datetime import date as date_type
            payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    plan_prices = {'basic': 20, 'pro': 50}
    per_month   = plan_prices.get(plan, 20)
    now_dt      = datetime.utcnow()
    months_label = f'{months} month' + ('' if months == 1 else 's')

    pr = PaymentRequest(
        user_id        = uid,
        plan           = plan,
        amount         = f'K{per_month}',
        months_paid    = months,
        payment_method = method,
        reference      = reference,
        notes          = notes,
        payment_date   = payment_date,
        status         = 'approved' if activate_now else 'pending_confirmation',
        submitted_at   = now_dt,
        resolved_at    = now_dt if activate_now else None,
    )
    db.session.add(pr)

    if activate_now:
        if target.plan == plan and target.plan_expires and target.plan_expires > now_dt:
            new_expires = target.plan_expires + timedelta(days=30 * months)
        else:
            target.plan_start = now_dt
            new_expires = now_dt + timedelta(days=30 * months)
        target.plan           = plan
        target.plan_expires   = new_expires
        target.monthly_builds_used = 0
        today = now_dt.date()
        target.monthly_reset_date = date(today.year + 1, 1, 1) if today.month == 12 \
                                    else date(today.year, today.month + 1, 1)
        if not target.catalogs:
            db.session.add(Catalog(user_id=target.id, name='My Catalog'))
        db.session.commit()
        log_admin_action(session['user_id'], 'payment_recorded_activated', 'payment_request', pr.id,
                         f'Recorded + activated {plan} × {months_label} for {target.name}')
        log_activity(target.id, 'plan_upgraded',
                     f'{plan.title()} plan activated ({months_label}) via admin-recorded payment')
        send_email(
            target.email,
            f'[CatalogKit] Your {plan.title()} plan is now active!',
            f"Hi {target.name},\n\n"
            f"We've confirmed your payment of K{per_month * months} ({months_label}).\n\n"
            f"Your {plan.title()} plan is now active and expires {new_expires.strftime('%d %b %Y')}.\n\n"
            f"Log in to start building your catalogs: https://www.catalogkit.org\n\n"
            f"Thank you!\n— The CatalogKit Team"
        )
        flash(f'Payment recorded and {plan.title()} plan activated for {target.name}.', 'success')
    else:
        db.session.commit()
        log_admin_action(session['user_id'], 'payment_logged', 'payment_request', pr.id,
                         f'Logged pending payment: {plan} × {months_label} for {target.name}')
        flash(f'Payment logged as "Pending Confirmation" for {target.name}.', 'success')

    return redirect(url_for('admin_billing'))


@app.route('/admin/payment/<int:pr_id>/confirm', methods=['POST'])
@admin_required
def admin_confirm_payment(pr_id):
    """Confirm receipt of a pending_confirmation payment and activate the plan."""
    pr = db.session.get(PaymentRequest, pr_id)
    if not pr or pr.status != 'pending_confirmation':
        flash('Payment record not found or already processed.', 'error')
        return redirect(url_for('admin_billing'))

    pr.status      = 'approved'
    pr.resolved_at = datetime.utcnow()

    user = db.session.get(User, pr.user_id)
    if user:
        months_paid  = pr.months_paid or 1
        now_dt       = datetime.utcnow()
        months_label = f'{months_paid} month' + ('' if months_paid == 1 else 's')
        if user.plan == pr.plan and user.plan_expires and user.plan_expires > now_dt:
            new_expires = user.plan_expires + timedelta(days=30 * months_paid)
        else:
            user.plan_start = now_dt
            new_expires = now_dt + timedelta(days=30 * months_paid)
        user.plan           = pr.plan
        user.plan_expires   = new_expires
        user.monthly_builds_used = 0
        today = now_dt.date()
        user.monthly_reset_date = date(today.year + 1, 1, 1) if today.month == 12 \
                                   else date(today.year, today.month + 1, 1)
        if not user.catalogs:
            db.session.add(Catalog(user_id=user.id, name='My Catalog'))
        db.session.commit()
        per_month_int = int(''.join(c for c in (pr.amount or '0') if c.isdigit()) or 0)
        total_kina    = per_month_int * months_paid
        log_admin_action(session['user_id'], 'payment_confirmed', 'payment_request', pr.id,
                         f'Confirmed receipt + activated {pr.plan} × {months_label} for {user.name}')
        log_activity(user.id, 'plan_upgraded',
                     f'{pr.plan.title()} plan activated ({months_label}) — receipt confirmed')
        send_email(
            user.email,
            f'[CatalogKit] Your {pr.plan.title()} plan is now active!',
            f"Hi {user.name},\n\n"
            f"We've confirmed your payment of K{total_kina} ({pr.amount}/month × {months_label}).\n\n"
            f"Your {pr.plan.title()} plan is now active and expires {new_expires.strftime('%d %b %Y')}.\n\n"
            f"Log in to start building your catalogs: https://www.catalogkit.org\n\n"
            f"Thank you!\n— The CatalogKit Team"
        )
        flash(f'Receipt confirmed — {pr.plan.title()} plan activated for {user.name}.', 'success')
    else:
        db.session.commit()
        flash('Payment confirmed (vendor account not found).', 'warning')

    return redirect(url_for('admin_billing'))


@app.route('/admin/payment/<int:pr_id>/delete-pending', methods=['POST'])
@admin_required
def admin_delete_pending_payment(pr_id):
    """Remove a pending_confirmation payment record (e.g. logged in error)."""
    pr = db.session.get(PaymentRequest, pr_id)
    if not pr or pr.status != 'pending_confirmation':
        flash('Record not found or already processed.', 'error')
        return redirect(url_for('admin_billing'))
    user = db.session.get(User, pr.user_id)
    log_admin_action(session['user_id'], 'payment_deleted', 'payment_request', pr.id,
                     f'Deleted pending payment record #{pr.id}' + (f' for {user.name}' if user else ''))
    db.session.delete(pr)
    db.session.commit()
    flash('Payment record deleted.', 'success')
    return redirect(url_for('admin_billing'))


# ── Done-For-You / Agency setup ───────────────────────────────────────────────

@app.route('/assisted-setup', methods=['GET', 'POST'])
def assisted_setup():
    if request.method == 'POST':
        business_name   = request.form.get('business_name', '').strip()
        market_location = request.form.get('market_location', '').strip()
        whatsapp        = request.form.get('whatsapp', '').strip()
        if not business_name or not market_location or not whatsapp:
            flash('All fields are required.', 'error')
            return render_template('assisted_setup.html', success=False)
        ar = AgencyRequest(business_name=business_name, market_location=market_location,
                           whatsapp=whatsapp, preferred_datetime='', catalog_plan='free')
        db.session.add(ar)
        db.session.commit()
        ae = admin_email()
        if ae:
            existing_acct = phone_in_use_by(whatsapp)
            dup_note = (
                f'\n⚠️  DUPLICATE WARNING: This WhatsApp is already linked to '
                f'"{existing_acct.name}" ({existing_acct.email}) — may be a re-submission.\n'
                if existing_acct else ''
            )
            send_email(ae,
                f'CatalogKit — New Done-For-You request: {business_name}',
                f'New Done-For-You setup request:\nName: {business_name}\n'
                f'Location: {market_location}\nWhatsApp: {whatsapp}'
                f'{dup_note}')
        return render_template('assisted_setup.html', success=True, ar=ar)
    return render_template('assisted_setup.html', success=False)


# ── Admin ─────────────────────────────────────────────────────────────────────

@app.route('/admin')
@admin_required
def admin():
    all_users        = User.query.order_by(User.created_at.desc()).all()
    agency_requests  = AgencyRequest.query.order_by(AgencyRequest.market_location, AgencyRequest.submitted_at).all()
    payment_requests = PaymentRequest.query.order_by(PaymentRequest.submitted_at.desc()).all()
    me = db.session.get(User, session['user_id'])
    pending_count = PaymentRequest.query.filter_by(status='pending').count()
    return render_template('admin.html', all_users=all_users,
                           agency_requests=agency_requests,
                           payment_requests=payment_requests,
                           pending_count=pending_count, now=datetime.utcnow(),
                           timedelta=timedelta,
                           me=me, admin_active='dashboard')

@app.route('/admin/agency/<int:ar_id>/status', methods=['POST'])
@admin_required
def update_agency_status(ar_id):
    ar = db.session.get(AgencyRequest, ar_id)
    if not ar:
        flash('Request not found.', 'error')
        return redirect(url_for('admin'))
    new_status = request.form.get('status', ar.status)
    allowed = ('New Request', 'Assigned', 'In Progress', 'Completed')
    linked_user_id = request.form.get('linked_user_id', '').strip()

    if new_status == 'Completed' and not linked_user_id and not ar.user_id:
        flash('Select the user account created for this business before marking the request as Completed.', 'error')
        return redirect(url_for('admin'))

    if new_status in allowed:
        if new_status == 'Completed' and ar.status != 'Completed':
            ar.completed_at = datetime.utcnow()
        elif new_status != 'Completed':
            ar.completed_at = None
        ar.status = new_status
        if linked_user_id:
            linked_user = db.session.get(User, int(linked_user_id))
            if linked_user:
                ar.user_id = linked_user.id
        db.session.commit()
        log_admin_action(session['user_id'], 'agency_status_changed', 'agency_request', ar.id,
                         f'{ar.business_name} → {new_status}')
    return redirect(url_for('admin'))

@app.route('/admin/user/<int:user_id>/suspend', methods=['POST'])
@full_admin_required
def suspend_user(user_id):
    user = db.session.get(User, user_id)
    if user and not user.is_admin:
        user.is_suspended = True
        user.suspended_at = datetime.utcnow()
        db.session.commit()
        log_admin_action(session['user_id'], 'user_suspended', 'user', user.id, f'{user.name} ({user.email})')
        # Log them out immediately if active
        flash(f'{user.name}\'s account has been suspended.', 'success')
        send_email(user.email,
            'CatalogKit — Your account has been suspended',
            f'Hi {user.name},\n\nYour CatalogKit account has been suspended. '
            f'Please contact us if you believe this is a mistake.'
        )
    return redirect(url_for('admin'))

@app.route('/admin/user/<int:user_id>/unsuspend', methods=['POST'])
@full_admin_required
def unsuspend_user(user_id):
    user = db.session.get(User, user_id)
    if user:
        user.is_suspended = False
        user.suspended_at = None
        db.session.commit()
        log_admin_action(session['user_id'], 'user_reinstated', 'user', user.id, f'{user.name} ({user.email})')
        flash(f'{user.name}\'s account has been reinstated.', 'success')
        send_email(user.email,
            'CatalogKit — Your account has been reinstated',
            f'Hi {user.name},\n\nGood news! Your CatalogKit account has been reinstated. '
            f'You can now log in and continue using CatalogKit.\n\n'
            f'Log in at: https://www.catalogkit.org'
        )
    return redirect(url_for('admin'))

@app.route('/admin/user/<int:user_id>/set-role', methods=['POST'])
@full_admin_required
def set_user_role(user_id):
    me = db.session.get(User, session['user_id'])
    target = db.session.get(User, user_id)
    if not target or target.id == me.id:
        flash('Cannot change your own role.', 'error')
        return redirect(url_for('admin'))
    role = request.form.get('role', 'user')
    target.is_admin     = (role == 'admin')
    target.is_moderator = (role == 'moderator')
    db.session.commit()
    log_activity(target.id, 'role_changed', f'Role set to {role} by admin {me.email}')
    log_admin_action(me.id, 'role_changed', 'user', target.id, f'{target.name} → {role}')
    flash(f'{target.name}\'s role updated to {role.title()}.', 'success')
    return redirect(url_for('admin'))

@app.route('/admin/user/create', methods=['POST'])
@full_admin_required
def admin_create_user():
    me       = db.session.get(User, session['user_id'])
    name     = request.form.get('name', '').strip()
    email    = request.form.get('email', '').strip().lower()
    password = request.form.get('password', '')
    access_level = request.form.get('access_level', 'user')
    is_tester    = request.form.get('is_tester') == 'on'
    if not name or not email or not password:
        flash('Name, email and password are required.', 'error')
        return redirect(url_for('admin'))
    if len(password) < 6:
        flash('Password must be at least 6 characters.', 'error')
        return redirect(url_for('admin'))
    if User.query.filter_by(email=email).first():
        flash('An account with that email already exists.', 'error')
        return redirect(url_for('admin'))
    user = User(name=name, email=email,
                password_hash=generate_password_hash(password),
                is_admin=(access_level == 'admin'),
                is_moderator=(access_level == 'moderator'),
                is_tester=is_tester)
    db.session.add(user)
    db.session.commit()
    log_activity(user.id, 'user_created_by_admin', f'Account created by admin {me.email}')
    log_admin_action(me.id, 'user_created', 'user', user.id, f'{user.name} ({user.email}) as {access_level}')
    flash(f'{user.name}\'s account has been created.', 'success')
    return redirect(url_for('admin'))

@app.route('/admin/user/<int:user_id>/edit', methods=['GET', 'POST'])
@full_admin_required
def admin_edit_user(user_id):
    me     = db.session.get(User, session['user_id'])
    target = db.session.get(User, user_id)
    if not target:
        flash('User not found.', 'error')
        return redirect(url_for('admin'))
    if request.method == 'POST':
        new_email = request.form.get('email', '').strip().lower()
        # Email uniqueness check (skip if unchanged)
        if new_email != target.email:
            existing = User.query.filter_by(email=new_email).first()
            if existing:
                flash('That email is already in use by another account.', 'error')
                return render_template('admin_edit_user.html', u=target, me=me,
                                       admin_active='users', now=datetime.utcnow(), pending_count=0)
        target.name           = request.form.get('name', target.name).strip()
        target.email          = new_email or target.email
        target.business_name  = request.form.get('business_name', '').strip() or None
        target.contact_person = request.form.get('contact_person', '').strip() or None
        target.location       = request.form.get('location', '').strip() or None
        target.whatsapp       = request.form.get('whatsapp', '').strip() or None
        target.phone          = request.form.get('phone', '').strip() or None
        new_password = request.form.get('new_password', '').strip()
        if new_password:
            target.password_hash = generate_password_hash(new_password)
        # Access level (only if editing someone other than yourself)
        if target.id != me.id:
            level = request.form.get('access_level', 'user')
            target.is_admin     = (level == 'admin')
            target.is_moderator = (level == 'moderator')
        target.is_tester = request.form.get('is_tester') == 'on'
        # Plan — lets the admin activate a plan directly (e.g. after a
        # WhatsApp payment) without needing a PaymentRequest record.
        new_plan = request.form.get('plan')
        plan_changed = new_plan in ('free', 'basic', 'pro') and new_plan != target.plan
        if plan_changed:
            old_plan = target.plan
            target.plan = new_plan
            target.monthly_builds_used = 0
            target.plan_start = datetime.utcnow()
            if not target.catalogs:
                catalog = Catalog(user_id=target.id, name='My Catalog')
                db.session.add(catalog)
            log_activity(target.id, 'plan_changed_by_admin',
                         f'Plan changed from {old_plan} to {new_plan} by admin {me.email}')
        db.session.commit()
        log_activity(target.id, 'profile_edited_by_admin', f'Profile edited by admin {me.email}')
        log_admin_action(me.id, 'user_edited', 'user', target.id, f'{target.name} ({target.email})')
        if plan_changed:
            send_email(
                target.email,
                f'[CatalogKit] Your {target.plan_label} plan is now active!',
                f"Hi {target.name},\n\n"
                f"Your {target.plan_label} plan is now active on CatalogKit.\n\n"
                f"Log in to start building your catalogs: https://www.catalogkit.org\n\n"
                f"Thank you for supporting CatalogKit!\n"
                f"— The CatalogKit Team"
            )
            flash(f'{target.name}\'s profile has been updated and their plan is now {target.plan_label}.', 'success')
        else:
            flash(f'{target.name}\'s profile has been updated.', 'success')
        return redirect(url_for('admin'))
    payment_records = PaymentRequest.query.filter_by(user_id=target.id)\
                      .filter(PaymentRequest.status.in_(['approved', 'pending_confirmation']))\
                      .order_by(PaymentRequest.submitted_at.desc()).all()
    return render_template('admin_edit_user.html', u=target, me=me,
                           admin_active='users', now=datetime.utcnow(), pending_count=0,
                           payment_records=payment_records)

@app.route('/admin/user/<int:user_id>/delete', methods=['POST'])
@full_admin_required
def admin_delete_user(user_id):
    me     = db.session.get(User, session['user_id'])
    target = db.session.get(User, user_id)
    if not target:
        flash('User not found.', 'error')
        return redirect(url_for('admin'))
    if target.id == me.id:
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('admin'))
    if target.is_admin:
        flash('Cannot delete another admin account.', 'error')
        return redirect(url_for('admin'))
    name = target.name
    email = target.email
    # Delete logs first to avoid NOT NULL constraint on user_id
    AccessLog.query.filter_by(user_id=target.id).delete()
    ActivityLog.query.filter_by(user_id=target.id).delete()
    SupportMessage.query.filter(SupportMessage.ticket_id.in_(
        db.session.query(SupportTicket.id).filter_by(user_id=target.id))).delete(synchronize_session=False)
    SupportTicket.query.filter_by(user_id=target.id).delete()
    db.session.delete(target)
    db.session.commit()
    log_admin_action(me.id, 'user_deleted', 'user', user_id, f'{name} ({email}) permanently deleted')
    flash(f'{name}\'s account and all their catalogs have been permanently deleted.', 'success')
    return redirect(url_for('admin'))

# ── Admin — Logs ───────────────────────────────────────────────────────────────

@app.route('/admin/logs')
@admin_required
def admin_logs():
    page      = int(request.args.get('page', 1))
    per_page  = 40
    log_type  = request.args.get('type', 'activity')   # 'activity' | 'access'
    user_id   = request.args.get('user_id', '', type=str)
    action_f  = request.args.get('action', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')

    all_users = User.query.order_by(User.name).all()

    if log_type == 'access':
        q = AccessLog.query
        if user_id:
            q = q.filter(AccessLog.user_id == int(user_id))
        if action_f:
            q = q.filter(AccessLog.action == action_f)
        if date_from:
            try: q = q.filter(AccessLog.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
            except ValueError: pass
        if date_to:
            try: q = q.filter(AccessLog.created_at <= datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
            except ValueError: pass
        logs      = q.order_by(AccessLog.created_at.desc()).offset((page-1)*per_page).limit(per_page).all()
        total     = q.count()
    else:
        q = ActivityLog.query
        if user_id:
            q = q.filter(ActivityLog.user_id == int(user_id))
        if action_f:
            q = q.filter(ActivityLog.action == action_f)
        if date_from:
            try: q = q.filter(ActivityLog.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
            except ValueError: pass
        if date_to:
            try: q = q.filter(ActivityLog.created_at <= datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
            except ValueError: pass
        logs      = q.order_by(ActivityLog.created_at.desc()).offset((page-1)*per_page).limit(per_page).all()
        total     = q.count()

    total_pages   = max(1, (total + per_page - 1) // per_page)
    pending_count = 0
    return render_template('admin_logs.html',
                           logs=logs, log_type=log_type, page=page,
                           total=total, total_pages=total_pages,
                           all_users=all_users, user_id=user_id,
                           action_f=action_f, date_from=date_from, date_to=date_to,
                           pending_count=pending_count, now=datetime.utcnow(),
                           admin_active='logs')


# ── Admin — Reports ────────────────────────────────────────────────────────────

@app.route('/admin/reports')
@admin_required
def admin_reports():
    from sqlalchemy import func as sqlfunc
    now   = datetime.utcnow()
    day30 = now - timedelta(days=30)
    day7  = now - timedelta(days=7)
    tab   = request.args.get('tab', 'overview')
    if tab not in ('overview', 'finance', 'access', 'catalogs'):
        tab = 'overview'

    ctx = dict(now=now, pending_count=0, admin_active='reports', tab=tab)

    if tab == 'overview':
        # ── Summary cards
        total_users    = User.query.count()
        active_30d     = db.session.query(AccessLog.user_id).filter(
                             AccessLog.action == 'login',
                             AccessLog.created_at >= day30).distinct().count()
        suspended_ct   = User.query.filter_by(is_suspended=True).count()
        new_users_30d  = User.query.filter(User.created_at >= day30).count()
        total_catalogs = Catalog.query.count()
        total_pdfs     = db.session.query(db.func.sum(Catalog.pdf_downloads)).scalar() or 0

        # ── New signups per week (last 8 weeks)
        signup_weeks = []
        for i in range(7, -1, -1):
            week_start = now - timedelta(days=(i+1)*7)
            week_end   = now - timedelta(days=i*7)
            count = User.query.filter(User.created_at >= week_start,
                                      User.created_at < week_end).count()
            label = week_start.strftime('%d %b')
            signup_weeks.append({'label': label, 'count': count})

        # ── Activity breakdown (last 30 days)
        activity_labels = {
            'catalog_created':   'Catalogs Created',
            'catalog_published': 'Catalogs Published',
            'images_uploaded':   'Image Uploads',
            'pdf_downloaded':    'PDFs Downloaded',
            'catalog_deleted':   'Catalogs Deleted',
            'catalog_renamed':   'Catalogs Renamed',
            'profile_updated':   'Profile Updates',
        }
        activity_counts = {}
        for key, label in activity_labels.items():
            activity_counts[label] = ActivityLog.query.filter(
                ActivityLog.action == key,
                ActivityLog.created_at >= day30).count()

        # ── Login activity (last 30 days by day)
        login_days = []
        for i in range(29, -1, -1):
            d = now - timedelta(days=i)
            d_start = d.replace(hour=0, minute=0, second=0, microsecond=0)
            d_end   = d_start + timedelta(days=1)
            count = AccessLog.query.filter(AccessLog.action == 'login',
                                           AccessLog.created_at >= d_start,
                                           AccessLog.created_at < d_end).count()
            login_days.append({'label': d.strftime('%d %b'), 'count': count})

        # ── Top 10 most active users (by total activity events)
        top_users = db.session.query(
            User, sqlfunc.count(ActivityLog.id).label('events')
        ).join(ActivityLog, ActivityLog.user_id == User.id)\
         .filter(ActivityLog.created_at >= day30)\
         .group_by(User.id).order_by(sqlfunc.count(ActivityLog.id).desc()).limit(10).all()

        ctx.update(total_users=total_users, active_30d=active_30d,
            suspended_ct=suspended_ct, new_users_30d=new_users_30d,
            total_catalogs=total_catalogs, total_pdfs=total_pdfs,
            signup_weeks=signup_weeks,
            activity_counts=activity_counts, login_days=login_days,
            top_users=top_users)

    elif tab == 'finance':
        plan_prices = {'free': 0, 'basic': 20, 'pro': 50}

        # ── Recurring plan revenue (based on current active paid users)
        plan_counts = {}
        for plan in ('free', 'basic', 'pro'):
            plan_counts[plan] = User.query.filter_by(plan=plan, is_admin=False, is_moderator=False).count()
        mrr = sum(plan_counts.get(p, 0) * price for p, price in plan_prices.items())

        # ── Payment requests (plan upgrades)
        pr_pending  = PaymentRequest.query.filter_by(status='pending').count()
        pr_approved = PaymentRequest.query.filter_by(status='approved').all()
        pr_rejected = PaymentRequest.query.filter_by(status='rejected').count()
        approved_total = sum(p.total_amount_kina for p in pr_approved)

        pr_by_method = {}
        for p in pr_approved:
            pr_by_method[p.payment_method] = pr_by_method.get(p.payment_method, 0) + 1

        # ── Active subscriptions & expiry tracking
        thirty_days_out = now + timedelta(days=30)
        active_basic   = User.query.filter(User.plan == 'basic',
                                           User.plan_expires > now).count()
        active_pro     = User.query.filter(User.plan == 'pro',
                                           User.plan_expires > now).count()
        expiring_soon  = User.query.filter(User.plan.in_(['basic', 'pro']),
                                           User.plan_expires > now,
                                           User.plan_expires <= thirty_days_out
                                           ).order_by(User.plan_expires).all()

        # ── Approved payments per week (last 8 weeks)
        revenue_weeks = []
        for i in range(7, -1, -1):
            week_start = now - timedelta(days=(i+1)*7)
            week_end   = now - timedelta(days=i*7)
            weekly = PaymentRequest.query.filter(
                PaymentRequest.status == 'approved',
                PaymentRequest.resolved_at >= week_start,
                PaymentRequest.resolved_at < week_end).all()
            total = sum(p.total_amount_kina for p in weekly)
            revenue_weeks.append({'label': week_start.strftime('%d %b'), 'count': total})

        # ── Done-For-You (agency) revenue: K50 visit fee + plan add-on
        agency_all = AgencyRequest.query.all()
        agency_completed = [a for a in agency_all if a.status == 'Completed']
        agency_revenue = sum(a.total_due for a in agency_completed)
        agency_by_plan = {}
        for a in agency_all:
            agency_by_plan[a.catalog_plan] = agency_by_plan.get(a.catalog_plan, 0) + 1

        recent_payments = PaymentRequest.query.order_by(PaymentRequest.submitted_at.desc()).limit(15).all()

        # ── Revenue by period (day / week / month / quarter / year) ──
        period = request.args.get('period', 'month')
        if period not in ('day', 'week', 'month', 'quarter', 'year'):
            period = 'month'

        def _bucket(dt):
            if period == 'day':
                return dt.strftime('%Y-%m-%d'), dt.strftime('%d %b %Y')
            if period == 'week':
                start = dt - timedelta(days=dt.weekday())
                return start.strftime('%Y-%m-%d'), 'Week of ' + start.strftime('%d %b %Y')
            if period == 'month':
                return dt.strftime('%Y-%m'), dt.strftime('%b %Y')
            if period == 'quarter':
                q = (dt.month - 1) // 3 + 1
                return f'{dt.year}-Q{q}', f'Q{q} {dt.year}'
            return str(dt.year), str(dt.year)

        buckets = {}
        for p in pr_approved:
            when = p.resolved_at or p.submitted_at
            if not when:
                continue
            key, label = _bucket(when)
            b = buckets.setdefault(key, {'label': label, 'plan_revenue': 0, 'agency_revenue': 0, 'tx_count': 0})
            b['plan_revenue'] += p.total_amount_kina
            b['tx_count'] += 1

        for a in agency_completed:
            when = a.completed_at or a.submitted_at
            if not when:
                continue
            key, label = _bucket(when)
            b = buckets.setdefault(key, {'label': label, 'plan_revenue': 0, 'agency_revenue': 0, 'tx_count': 0})
            b['agency_revenue'] += a.total_due
            b['tx_count'] += 1

        period_rows = []
        for key in sorted(buckets.keys(), reverse=True):
            b = buckets[key]
            period_rows.append({
                'label': b['label'],
                'plan_revenue': b['plan_revenue'],
                'agency_revenue': b['agency_revenue'],
                'total': b['plan_revenue'] + b['agency_revenue'],
                'tx_count': b['tx_count'],
            })
        period_total = sum(r['total'] for r in period_rows)
        period_labels = {'day': 'Day', 'week': 'Week', 'month': 'Month', 'quarter': 'Quarter', 'year': 'Year'}

        ctx.update(plan_counts=plan_counts, plan_prices=plan_prices, mrr=mrr,
            pr_pending=pr_pending, pr_approved_count=len(pr_approved), pr_rejected=pr_rejected,
            approved_total=approved_total, pr_by_method=pr_by_method, revenue_weeks=revenue_weeks,
            period=period, period_label=period_labels[period], period_rows=period_rows, period_total=period_total,
            agency_total=len(agency_all), agency_completed_count=len(agency_completed),
            agency_revenue=agency_revenue, agency_by_plan=agency_by_plan,
            recent_payments=recent_payments,
            active_basic=active_basic, active_pro=active_pro,
            expiring_soon=expiring_soon)

    elif tab == 'access':
        total_users        = User.query.count()
        admins             = User.query.filter_by(is_admin=True).order_by(User.name).all()
        moderators         = User.query.filter_by(is_moderator=True).order_by(User.name).all()
        regular_users      = User.query.filter_by(is_admin=False, is_moderator=False).count()
        regular_users_list = User.query.filter_by(is_admin=False, is_moderator=False)\
                                 .order_by(User.created_at.desc()).all()
        testers            = User.query.filter_by(is_tester=True).count()
        tester_users_list  = User.query.filter_by(is_tester=True).order_by(User.name).all()
        suspended          = User.query.filter_by(is_suspended=True).all()

        plan_breakdown = {}
        for plan in ('free', 'basic', 'pro'):
            plan_breakdown[plan] = User.query.filter_by(plan=plan).count()

        never_logged_in = User.query.outerjoin(
            AccessLog, db.and_(AccessLog.user_id == User.id, AccessLog.action == 'login')
        ).filter(AccessLog.id.is_(None)).count()

        active_ids_30d = {uid for (uid,) in db.session.query(AccessLog.user_id).filter(
            AccessLog.action == 'login', AccessLog.created_at >= day30).distinct()}
        dormant_30d = max(0, total_users - len(active_ids_30d))

        recent_access = AccessLog.query.order_by(AccessLog.created_at.desc()).limit(20).all()

        # ── Duplicate phone/WhatsApp detection ─────────────────────────────
        # Group users by their normalized phone number; flag groups with >1 account.
        phone_groups = {}
        all_users_for_dup = User.query.filter_by(is_admin=False, is_moderator=False).all()
        for u in all_users_for_dup:
            for raw in filter(None, [u.whatsapp, u.phone]):
                norm = normalize_phone(raw)
                if norm:
                    phone_groups.setdefault(norm, [])
                    if not any(x.id == u.id for x in phone_groups[norm]):
                        phone_groups[norm].append(u)
        duplicate_groups = [
            {'phone': norm, 'users': users}
            for norm, users in phone_groups.items()
            if len(users) > 1
        ]

        ctx.update(total_users=total_users, admins=admins, moderators=moderators,
            regular_users=regular_users, regular_users_list=regular_users_list,
            testers=testers, tester_users_list=tester_users_list, suspended=suspended,
            plan_breakdown=plan_breakdown, never_logged_in=never_logged_in,
            dormant_30d=dormant_30d, recent_access=recent_access,
            duplicate_groups=duplicate_groups)

    elif tab == 'catalogs':
        total_catalogs     = Catalog.query.count()
        published_ct       = Catalog.query.filter_by(is_published=True).count()
        draft_ct           = total_catalogs - published_ct
        total_pdfs         = db.session.query(db.func.sum(Catalog.pdf_downloads)).scalar() or 0
        avg_pages          = db.session.query(sqlfunc.avg(Catalog.page_count)).scalar() or 0
        empty_catalogs     = Catalog.query.filter(Catalog.page_count == 0).count()

        layout_counts = {}
        for row in db.session.query(User.pdf_layout, sqlfunc.count(Catalog.id))\
                .join(Catalog, Catalog.user_id == User.id)\
                .group_by(User.pdf_layout).all():
            layout_counts[row[0] or 'classic'] = row[1]

        top_catalogs = Catalog.query.filter(Catalog.pdf_downloads > 0)\
            .order_by(Catalog.pdf_downloads.desc()).limit(10).all()

        agency_by_status = {}
        for row in db.session.query(AgencyRequest.status, sqlfunc.count(AgencyRequest.id))\
                .group_by(AgencyRequest.status).all():
            agency_by_status[row[0] or 'New Request'] = row[1]

        ctx.update(total_catalogs=total_catalogs, published_ct=published_ct, draft_ct=draft_ct,
            total_pdfs=total_pdfs, avg_pages=round(avg_pages, 1), empty_catalogs=empty_catalogs,
            layout_counts=layout_counts, top_catalogs=top_catalogs, agency_by_status=agency_by_status)

    return render_template('admin_reports.html', **ctx)


# ── Admin — Audit Log (who changed what) ────────────────────────────────────────

@app.route('/admin/audit-log')
@full_admin_required
def admin_audit_log():
    page      = int(request.args.get('page', 1))
    per_page  = 40
    admin_id  = request.args.get('admin_id', '', type=str)
    action_f  = request.args.get('action', '')

    q = AdminAuditLog.query
    if admin_id:
        q = q.filter(AdminAuditLog.admin_id == int(admin_id))
    if action_f:
        q = q.filter(AdminAuditLog.action == action_f)

    total       = q.count()
    total_pages = max(1, (total + per_page - 1) // per_page)
    entries     = q.order_by(AdminAuditLog.created_at.desc()).offset((page-1)*per_page).limit(per_page).all()
    admins      = User.query.filter(db.or_(User.is_admin == True, User.is_moderator == True)).order_by(User.name).all()
    distinct_actions = [r[0] for r in db.session.query(AdminAuditLog.action).distinct().order_by(AdminAuditLog.action).all()]

    return render_template('admin_audit_log.html', entries=entries, page=page, total=total,
                           total_pages=total_pages, admins=admins, admin_id=admin_id,
                           action_f=action_f, distinct_actions=distinct_actions,
                           pending_count=0, now=datetime.utcnow(), admin_active='audit')


# ── Admin — Announcements (bulk vendor notifications) ───────────────────────────

@app.route('/admin/announcements', methods=['GET', 'POST'])
@full_admin_required
def admin_announcements():
    me = db.session.get(User, session['user_id'])
    if request.method == 'POST':
        subject  = request.form.get('subject', '').strip()
        message  = request.form.get('message', '').strip()
        audience = request.form.get('audience', 'all')
        if not subject or not message:
            flash('Please fill in both a subject and a message.', 'error')
            return redirect(url_for('admin_announcements'))

        q = User.query.filter_by(is_suspended=False)
        if audience == 'free':
            q = q.filter_by(plan='free', is_admin=False, is_moderator=False)
        elif audience == 'basic':
            q = q.filter_by(plan='basic')
        elif audience == 'pro':
            q = q.filter_by(plan='pro')
        elif audience == 'paid':
            q = q.filter(User.plan.in_(['basic', 'pro']))
        else:
            q = q.filter_by(is_admin=False, is_moderator=False)
        recipients = q.all()

        sent = 0
        for u in recipients:
            body = f"Hi {u.name},\n\n{message}\n\n— The CatalogKit Team"
            if send_email(u.email, f'[CatalogKit] {subject}', body):
                sent += 1

        log_admin_action(me.id, 'announcement_sent', 'broadcast', None,
                         f'"{subject}" to {audience} ({sent}/{len(recipients)} delivered)')
        if sent == 0 and recipients:
            flash('Message could not be sent — check that email sending (MAIL_USERNAME/MAIL_PASSWORD) is configured.', 'error')
        else:
            flash(f'Announcement sent to {sent} of {len(recipients)} vendor(s).', 'success')
        return redirect(url_for('admin_announcements'))

    recent = AdminAuditLog.query.filter_by(action='announcement_sent')\
                                 .order_by(AdminAuditLog.created_at.desc()).limit(10).all()
    audience_counts = {
        'all':   User.query.filter_by(is_suspended=False, is_admin=False, is_moderator=False).count(),
        'free':  User.query.filter_by(is_suspended=False, plan='free', is_admin=False, is_moderator=False).count(),
        'basic': User.query.filter_by(is_suspended=False, plan='basic').count(),
        'pro':   User.query.filter_by(is_suspended=False, plan='pro').count(),
        'paid':  User.query.filter(User.is_suspended == False, User.plan.in_(['basic', 'pro'])).count(),
    }
    return render_template('admin_announcements.html', me=me, recent=recent, audience_counts=audience_counts,
                           pending_count=0, now=datetime.utcnow(), admin_active='announcements')


# ── Admin — Support inbox ────────────────────────────────────────────────────────

@app.route('/admin/support')
@admin_required
def admin_support():
    status_f = request.args.get('status', 'open')
    q = SupportTicket.query
    if status_f in ('open', 'in_progress', 'resolved'):
        q = q.filter_by(status=status_f)
    tickets    = q.order_by(SupportTicket.updated_at.desc()).all()
    open_count = SupportTicket.query.filter_by(status='open').count()
    return render_template('admin_support.html', tickets=tickets, status_f=status_f,
                           open_count=open_count, pending_count=0, now=datetime.utcnow(),
                           admin_active='support')

@app.route('/admin/support/<int:ticket_id>', methods=['GET', 'POST'])
@admin_required
def admin_support_ticket(ticket_id):
    me     = db.session.get(User, session['user_id'])
    ticket = db.session.get(SupportTicket, ticket_id)
    if not ticket:
        flash('Ticket not found.', 'error')
        return redirect(url_for('admin_support'))
    if request.method == 'POST':
        reply      = request.form.get('reply', '').strip()
        new_status = request.form.get('status', ticket.status)
        if reply:
            db.session.add(SupportMessage(ticket_id=ticket.id, from_admin=True,
                                          author_name=me.name, body=reply))
            send_email(ticket.user.email, f'[CatalogKit Support] Re: {ticket.subject}',
                      f"Hi {ticket.user.name},\n\n{reply}\n\n— The CatalogKit Team\n\n"
                      f"View this conversation any time at: https://www.catalogkit.org/support/{ticket.id}")
        if new_status in ('open', 'in_progress', 'resolved'):
            ticket.status = new_status
        ticket.updated_at = datetime.utcnow()
        db.session.commit()
        log_admin_action(me.id, 'support_replied', 'support_ticket', ticket.id,
                         f'Re: {ticket.subject} (status: {ticket.status})')
        flash('Reply sent.', 'success')
        return redirect(url_for('admin_support_ticket', ticket_id=ticket.id))
    return render_template('admin_support_ticket.html', ticket=ticket, pending_count=0,
                           now=datetime.utcnow(), admin_active='support')


# ── Admin — Excel exports ─────────────────────────────────────────────────────────

def build_xlsx_response(filename, headers, rows):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for row in rows:
        ws.append(row)
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(str(header)) + 4)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/export/users.xlsx')
@admin_required
def export_users_xlsx():
    headers = ['Name', 'Email', 'Business', 'WhatsApp', 'Role', 'Plan', 'Catalogs', 'Status', 'Joined']
    rows = []
    for u in User.query.order_by(User.created_at.desc()).all():
        role = 'Admin' if u.is_admin else ('Moderator' if u.is_moderator else 'User')
        rows.append([u.name, u.email, u.business_name or '', u.whatsapp or '', role,
                    u.plan_label, u.catalog_count, 'Suspended' if u.is_suspended else 'Active',
                    u.created_at.strftime('%Y-%m-%d')])
    log_admin_action(session['user_id'], 'export_users', detail=f'{len(rows)} users exported to Excel')
    return build_xlsx_response('catalogkit-users.xlsx', headers, rows)

@app.route('/admin/export/finance.xlsx')
@admin_required
def export_finance_xlsx():
    headers = [
        'Date Approved', 'Type', 'Vendor Name', 'Email', 'Business',
        'Plan', 'Rate (K/mo)', 'Months Paid', 'Total (K)',
        'Method', 'Reference', 'Status',
        'Plan Start', 'Plan Expires', 'Next Payment Due',
    ]
    rows = []
    now_dt = datetime.utcnow()
    for p in PaymentRequest.query.order_by(PaymentRequest.submitted_at.desc()).all():
        u = p.user
        resolved = p.resolved_at or p.submitted_at
        plan_expires = u.plan_expires.strftime('%Y-%m-%d') if u and u.plan_expires else ''
        plan_start   = u.plan_start.strftime('%Y-%m-%d')   if u and u.plan_start   else ''
        next_due = plan_expires  # For manual billing, next_due == expiry date
        rows.append([
            resolved.strftime('%Y-%m-%d'),
            'Plan Payment',
            u.name         if u else '',
            u.email        if u else '',
            u.business_name if u and u.business_name else '',
            p.plan,
            p.amount,
            p.months_paid or 1,
            p.total_amount_kina,
            p.payment_method,
            p.reference or '',
            p.status,
            plan_start,
            plan_expires,
            next_due,
        ])
    for a in AgencyRequest.query.order_by(AgencyRequest.submitted_at.desc()).all():
        rows.append([
            a.submitted_at.strftime('%Y-%m-%d'), 'Done-For-You',
            a.business_name, '', '',
            a.catalog_plan, '', 1, a.total_due,
            'Agent visit', '', a.status, '', '', '',
        ])
    log_admin_action(session['user_id'], 'export_finance', detail=f'{len(rows)} rows exported to Excel')
    return build_xlsx_response('catalogkit-finance.xlsx', headers, rows)


# ── Pricing page ──────────────────────────────────────────────────────────────

@app.route('/pricing')
def pricing():
    return render_template('pricing.html')


# ── Contact & FAQ page ──────────────────────────────────────────────────────────

@app.route('/contact')
def contact():
    return render_template('contact.html')


# ── Support inbox (vendor-facing) ────────────────────────────────────────────────

def admin_email():
    admin = User.query.filter_by(is_admin=True).order_by(User.id).first()
    return admin.email if admin else None

@app.route('/support', methods=['GET', 'POST'])
@login_required
def support():
    user = current_user()
    if request.method == 'POST':
        subject = request.form.get('subject', '').strip()
        message = request.form.get('message', '').strip()
        if not subject or not message:
            flash('Please fill in both a subject and your message.', 'error')
            return redirect(url_for('support'))
        ticket = SupportTicket(user_id=user.id, subject=subject)
        db.session.add(ticket)
        db.session.commit()
        db.session.add(SupportMessage(ticket_id=ticket.id, from_admin=False,
                                      author_name=user.name, body=message))
        db.session.commit()
        log_activity(user.id, 'support_ticket_created', subject[:200])
        ae = admin_email()
        if ae:
            send_email(ae, f'[CatalogKit Support] New message from {user.name}: {subject}',
                      f"{user.name} ({user.email}) sent a support message:\n\n"
                      f"Subject: {subject}\n\n{message}\n\n"
                      f"Reply from the admin panel: https://www.catalogkit.org/admin/support/{ticket.id}")
        flash("Your message has been sent — we'll reply here and by email.", 'success')
        return redirect(url_for('support_ticket', ticket_id=ticket.id))
    tickets = SupportTicket.query.filter_by(user_id=user.id).order_by(SupportTicket.created_at.desc()).all()
    return render_template('support.html', user=user, tickets=tickets)

@app.route('/support/<int:ticket_id>', methods=['GET', 'POST'])
@login_required
def support_ticket(ticket_id):
    user = current_user()
    ticket = db.session.get(SupportTicket, ticket_id)
    if not ticket or ticket.user_id != user.id:
        flash('That message thread was not found.', 'error')
        return redirect(url_for('support'))
    if request.method == 'POST':
        body = request.form.get('body', '').strip()
        if body:
            db.session.add(SupportMessage(ticket_id=ticket.id, from_admin=False,
                                          author_name=user.name, body=body))
            if ticket.status == 'resolved':
                ticket.status = 'open'
            ticket.updated_at = datetime.utcnow()
            db.session.commit()
            ae = admin_email()
            if ae:
                send_email(ae, f'[CatalogKit Support] {user.name} replied: {ticket.subject}',
                          f"{user.name} ({user.email}) replied:\n\n{body}\n\n"
                          f"Reply from the admin panel: https://www.catalogkit.org/admin/support/{ticket.id}")
        return redirect(url_for('support_ticket', ticket_id=ticket.id))
    return render_template('support_ticket.html', user=user, ticket=ticket)


# ── Context processor ─────────────────────────────────────────────────────────

@app.context_processor
def inject_globals():
    cu = current_user()
    ctx = dict(current_user=cu, now=datetime.utcnow())
    if cu and (cu.is_admin or cu.is_moderator):
        ctx['open_support_count'] = SupportTicket.query.filter_by(status='open').count()
    return ctx


# ── Startup: create tables + run column migrations ────────────────────────────

with app.app_context():
    db.create_all()
    _migrations = [
        # ── user: profile fields ───────────────────────────────────────────────
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS contact_person VARCHAR(255)',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS phone VARCHAR(50)',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS facebook_url VARCHAR(500)',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS payment_methods TEXT',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS bank_account_details TEXT',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS delivery_methods TEXT',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS logo_filename VARCHAR(500)',
        "ALTER TABLE \"user\" ADD COLUMN IF NOT EXISTS brand_color VARCHAR(7)",
        "ALTER TABLE \"user\" ADD COLUMN IF NOT EXISTS pdf_layout VARCHAR(20) DEFAULT 'classic'",
        # ── user: auth / password reset ────────────────────────────────────────
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS reset_token VARCHAR(100)',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS reset_token_exp TIMESTAMP',
        # ── user: plan tracking ────────────────────────────────────────────────
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS plan_expires TIMESTAMP',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS plan_start TIMESTAMP',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS monthly_builds_used INTEGER DEFAULT 0',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS monthly_reset_date DATE',
        # ── user: access control ───────────────────────────────────────────────
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS is_moderator BOOLEAN DEFAULT FALSE',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS is_tester BOOLEAN DEFAULT FALSE',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS is_suspended BOOLEAN DEFAULT FALSE',
        'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS suspended_at TIMESTAMP',
        # ── catalog ────────────────────────────────────────────────────────────
        'ALTER TABLE catalog ADD COLUMN IF NOT EXISTS pdf_downloads INTEGER DEFAULT 0',
        'ALTER TABLE catalog ADD COLUMN IF NOT EXISTS is_published BOOLEAN DEFAULT FALSE',
        # ── agency_request ─────────────────────────────────────────────────────
        "ALTER TABLE agency_request ADD COLUMN IF NOT EXISTS catalog_plan VARCHAR(20) DEFAULT 'free'",
        'ALTER TABLE agency_request ADD COLUMN IF NOT EXISTS user_id INTEGER REFERENCES "user"(id)',
        'ALTER TABLE agency_request ADD COLUMN IF NOT EXISTS completed_at TIMESTAMP',
        '''CREATE TABLE IF NOT EXISTS payment_request (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES "user"(id),
            plan VARCHAR(20) NOT NULL,
            amount VARCHAR(10) NOT NULL,
            months_paid INTEGER DEFAULT 1,
            payment_method VARCHAR(50) NOT NULL,
            reference VARCHAR(255),
            status VARCHAR(20) DEFAULT \'pending\',
            notes VARCHAR(500),
            submitted_at TIMESTAMP DEFAULT NOW(),
            resolved_at TIMESTAMP
        )''',
        'ALTER TABLE payment_request ADD COLUMN IF NOT EXISTS months_paid INTEGER DEFAULT 1',
    ]
    with db.engine.connect() as _conn:
        for _sql in _migrations:
            try:
                _conn.execute(text(_sql))
                _conn.commit()
            except Exception as _e:
                app.logger.warning('Migration skipped: %s', _e)

# ── About page ─────────────────────────────────────────────────────────────────

@app.route('/about')
def about():
    return render_template('about.html')


# ── Privacy Policy ──────────────────────────────────────────────────────────────

@app.route('/privacy')
def privacy():
    return render_template('privacy.html')


# ── Terms of Service ────────────────────────────────────────────────────────────

@app.route('/terms')
def terms():
    return render_template('terms.html')


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8000))
    app.run(host='0.0.0.0', port=port, debug=False)
